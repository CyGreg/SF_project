import csv
import sys
import datetime
import openpyxl
from datetime import date, time
import gurobipy as gp
from gurobipy import *
import holidays
import calendar

print('wersja 4 nowa')
pathParam = 'anxDailyParametry.xlsx'

TODAY = datetime.datetime.now().date()

class ZLECENIE:
    def __init__(self, nbr, material, date_r, weight, originalWeight, group, divisible, indSubset, prefLine, priority, materialName):
        self.nbr = nbr
        self.material = material
        self.date_r = date_r
        self.weight = weight
        self.originalWeight = originalWeight
        self.group = group
        self.divisible = divisible
        self.indSubset = indSubset
        self.prefLine = prefLine
        self.priority = priority
        self.materialName = materialName

class WYNIK:
    def __init__(self, line, nbr, material, weight, originalWeight, time, order, realTimeStart, realTimeStop, materialName, prefLine, group, date_r, indSubset):
        self.line = line
        self.nbr = nbr
        self.material = material
        self.weight = weight
        self.originalWeight = originalWeight
        self.time = time
        self.order = order
        self.realTimeStart = realTimeStart
        self.realTimeStop = realTimeStop
        self.materialName = materialName
        self.prefLine = prefLine
        self.group = group
        self.date_r = date_r
        self.indSubset = indSubset

class MIX_PROD:
    def __init__(self, mat, lpr1, lpr2, udz1, udz2):
        self.material = mat
        self.lPrio1 = lpr1
        self.lPrio2 = lpr2
        self.udzial1 = udz1
        self.udzial2 = udz2

def load_data_from_excel(wb_name,sh_name,r0,c0):      # zwraca wiersze excela jako listę słowników (klucze każdego słownika to pierwszy wiersz excela)
    wb = openpyxl.load_workbook(wb_name, data_only=True)
    
    sh = wb[sh_name]
    min_row = r0
    min_col = c0
    max_row = sh.max_row
    max_col = sh.max_column
    fields = []

    for j in range(min_col,max_col+1):
        fields.append(sh.cell(row = min_row,column = j).value)

    rows = []
    for i in range(min_row + 1, max_row + 1):
        row = {}

        for j in range(min_col, max_col + 1):
            row[fields[j - 1]] = sh.cell(row=i, column=j).value
        rows.append(row)

    return rows

def load_MP(wb_name,group,sh_names,gCData):
    mp = {}
    mp_values = {}

    wb = openpyxl.load_workbook(wb_name, data_only=True)

    rows = []
    for sh_name in sh_names:
        try:
            sh = wb[sh_name]
        except KeyError:
            print('WARNING: brak przezbrojeń dla linii', sh_name)
            continue

        max_row = sh.max_row
        max_col = sh.max_column
        
        if sh_name in ['L1','L10','L7','L24','L4','L5','L23']:
            def_row = 3
            min_row = 5
            def_col = 2
            min_col = 5
        elif sh_name in ['L3','L16','L20','L8','L2','wybijanie logów','OCM','VAC']:
            def_row = 4
            min_row = 5
            def_col = 2
            min_col = 5
        
        fields = ['p1','p2','linia','mp(p1,p2,l)']

        for i in range(min_row, max_row + 1):
            p1 = str(sh.cell(row = i, column = def_col).value)
            if len(p1) < 3 or p1 in [None,'None']:
                continue
            for j in range(min_col, max_col + 1):
                p2 = str(sh.cell(row = def_row, column = j).value)
                if len(p2) < 3 or p2 in [None,'None']:
                    continue
                mp[(p1,p2,sh_name)] = sh.cell(row = i, column = j).value
                if mp[(p1,p2,sh_name)] == None:
                    mp[(p1, p2, sh_name)] = 0
                try:
                    if group[p1] not in [None,'None'] and group[p2] not in [None,'None']:
                        mp[(p1, p2, sh_name)] += gCData[(group[p1],group[p2])]
                except KeyError:
                    pass

                rows.append([p1,p2,sh_name,mp[(p1,p2,sh_name)]])
                try:
                    if mp[(p1,p2,sh_name)] not in mp_values[sh_name]:
                        mp_values[sh_name].append(mp[(p1,p2,sh_name)])
                except KeyError:
                    mp_values[sh_name] = [ mp[(p1, p2, sh_name)] ]

    with open('ANX_MP.csv', 'w', errors="ignore", encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(fields)
        write.writerows(rows)

    return mp

def load_efficiency(wb_name, sh_name):
    lines = []
    products = []
    eff = {}
    nTime = {}
    group = {}
    prLines = {}

    wb = openpyxl.load_workbook(wb_name, data_only=True)
    sh = wb[sh_name]

    licznik = 2
    while True:
        licznik += 1
        prod = str(sh.cell(row=licznik, column = 17).value)
        if prod in [None,'None','']:
            break
        if prod not in products:
            products.append(prod)

        group[prod] = str(sh.cell(row = licznik, column = 22).value)
        
        prLines[prod] = str(sh.cell(row = licznik, column = 39).value)

        for j in range(23,38):
            line = str(sh.cell(row=2,column=j).value)
            if line not in lines:
                lines.append(line)

            effka = sh.cell(row=licznik,column=j).value

            if effka not in [0, None]:
                eff[(prod,line)] = effka # ile kg na h
                nTime[(prod,line)] = 60 / effka # ile h na kg * 60 = ile min na kg

    rows = []
    for effka in eff:
        rows.append([effka,eff[effka]])
        
    with open('ANX_PRODUCTS_EFFICIENCY.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['produkt', 'wydajnosci'])
        write.writerows(rows)

    return eff, lines, group, products, prLines, nTime

def loadParameters(wb_name,sh_name,nCols):
    
    wb = openpyxl.load_workbook(wb_name, data_only=True)
    sh = wb[sh_name]
    param = {}
    
    #nCols 1: 1 kolumna klucz, 1 kolumna wartość; 2: 2 kolumny klucz, 1 kolumna wartość;
    # 3: 1 kolumna klucz, 2 i 3 kolumna wartości z drugim kluczem w nagłówku;
    # 0: wartości z jedynej kolumny, wchodzą do listy
    
    if nCols == 1:
        licznik = 1
        while True:
            licznik += 1
            nKey = str(sh.cell(row=licznik, column = 1).value)
            if nKey in (None,'None',''):
                break
            else:
                try:
                    nVal = float(sh.cell(row=licznik, column = 2).value)
                except:
                    nVal = sh.cell(row=licznik, column = 2).value
                finally:
                    param[nKey] = nVal
        return param
    
    elif nCols == 2:
        licznik = 1
        while True:
            licznik += 1
            nKey = str(sh.cell(row=licznik, column = 1).value)
            nKey2 = str(sh.cell(row=licznik, column = 2).value)
            if nKey in (None,'None',''):
                break
            else:
                try:
                    nVal = float(sh.cell(row=licznik, column = 3).value)
                except:
                    nVal = sh.cell(row=licznik, column = 3).value
                finally:
                    param[nKey,nKey2] = nVal
        return param
    
    elif nCols >= 3:
        nTitle=[]
        nrCol=2
        nColKey=str(sh.cell(row=1, column = nrCol).value)
        while nColKey not in (None,'None',''):
            nTitle.append(nColKey)
            nrCol += 1
            nColKey=str(sh.cell(row=1, column = nrCol).value)
        
        keylist = []
        licznik = 1
        while True:
            licznik += 1
            nKey = str(sh.cell(row=licznik, column = 1).value)

            if nKey in (None,'None',''):
                break
            else:
                for nrCol in range (2,nCols + 1):
                    nVal = float(sh.cell(row=licznik, column = nrCol).value)
                    param[nKey,nTitle[nrCol-2]] = nVal
                keylist.append(nKey)
        return param, keylist    

def edit_sales_orders(sales,group,subset,prefLines,wTermin,wIlosc,tps_list):
    items = []

    rows = []
    
    if subset == 'J2':
        indJ2 = []
        indJ2_1 = []
        indJ2_2 = []
    elif subset == 'J3':
        indJ3 = []
        produkcyjne = [] #alert AZ

        for zlec in ITEMS_PROD: #alert AZ
            if zlec.material not in produkcyjne:#alert AZ
                produkcyjne.append(zlec.material)#alert AZ
    
    for row in sales:
        date_r = row['Bazowy termin rozp.']
        material = str(row['Nr materiału'])
        matName = str(row['Krótki tekst materiału'])
        nbr = str(row['Nr zlecenia'])
        if nbr in ['None','']:
            continue
            
        if subset == 'J3' and material not in produkcyjne and material in matIndxMix: #alert AZ
            continue  # alert AZ
            
        if subset == 'J3':
            strDate = date_r
            someday = datetime.date(int(strDate[6:10]),int(strDate[3:5]),int(strDate[0:2]))
            dds = float((someday - TODAY).days)
            try:
                teps = tps_list[material]
            except:
                teps = 0
            if teps < dds or dds > PARAM_GLOBAL['daysPlan']:
                continue
                
        try:
            weight = float(row['waga'])
        except:
            print('WARNING: błędne dane dla zlecenia (brak wagi?)', nbr)
            weight=float(-1)
            continue
        grupa = group.get(material,'brak')
        
        
        
        if subset == 'J2':
            if material in matIndxMix:
                indJ2.append(nbr+'m1')
                indJ2.append(nbr+'m2')
            else:
                indJ2.append(nbr)
            prior = 0
            podz = row['opóźnienie'] 
            if podz == 0: #niepodzielny
                nSubset = 'J2_1'
                if material in matIndxMix:
                    indJ2_1.append(nbr+'m1')
                    indJ2_1.append(nbr+'m2')
                else:
                    indJ2_1.append(nbr)
            else: #podzielny
                nSubset = 'J2_2'
                if material in matIndxMix:
                    indJ2_2.append(nbr+'m1')
                    indJ2_2.append(nbr+'m2')
                else:
                    indJ2_2.append(nbr)
        elif subset == 'J3':
            prior = 666 #wartość będzie później korygowana
            podz = 1
            nSubset = subset
            if material in matIndxMix:
                indJ3.append(nbr+'m1')
                indJ3.append(nbr+'m2')
            else:
                indJ3.append(nbr)
        
        try:
            nPrefLine = prefLines[material]
        except:
            nPrefLine = 'brak'
            print('WARNING: Niezdefiniowana linia preferowana dla indx ',material, 'przyjmuję brak')
        finally:
            if material in matIndxMix:
                wgt1 = weight * MAT_MIX[matIndxMix.index(material)].udzial1
                wgt2 = weight * MAT_MIX[matIndxMix.index(material)].udzial2
                lpi1 =  MAT_MIX[matIndxMix.index(material)].lPrio1
                lpi2 =  MAT_MIX[matIndxMix.index(material)].lPrio2
                items.append(ZLECENIE(nbr+'m1',material,date_r,wgt1,wgt1,grupa,podz,nSubset,lpi1,prior,matName))
                items.append(ZLECENIE(nbr+'m2',material,date_r,wgt2,wgt2,grupa,podz,nSubset,lpi2,prior,matName))
                item = items[-2]
                rows.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset, item.prefLine,item.priority])
                item = items[-1]
                rows.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset, item.prefLine,item.priority])
            else:
                items.append(ZLECENIE(nbr,material,date_r,weight,weight,grupa,podz,nSubset,nPrefLine,prior,matName))
                item = items[-1]
                rows.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset, item.prefLine,item.priority])
              

    #przeliczenie priorytetów dla planu        
    if subset == 'J3':
        prio1 = []
        prio2 = []
        prioS1 = []
        prioS2 = []
        minT = float(999999)
        minW = float(999999)
        maxT = float(0)
        maxW = float(0)
        
        for itNr in range(0,len(items)):
            strDate = items[itNr].date_r
            someday = datetime.date(int(strDate[6:10]),int(strDate[3:5]),int(strDate[0:2]))
            dds = float((someday - TODAY).days)
            if dds < 0:
                dds = float(0)
            prio1.append(dds)
            prioS1.append(float(1))
            if minT > dds:
                minT = dds
            if maxT < dds:
                maxT = dds
            wgt = float(items[itNr].weight)
            if wgt < 0:
                wgt = float(0)
            prio2.append(wgt)
            prioS2.append(float(1))
            if minW > wgt:
                minW = wgt
            if maxW < wgt:
                maxW = wgt
        
        if minT != maxT:
            for itNr in range(0,len(items)):
                prioS1[itNr] = (maxT - prio1[itNr]) / (maxT - minT) 
                
        if minW != maxW:
            for itNr in range(0,len(items)):
                prioS2[itNr] = (prio2[itNr] - minW) / (maxW - minW)
                
        
        for itNr in range(0,len(items)):
            items[itNr].priority = wTermin * prioS1[itNr] + wIlosc * prioS2[itNr]
            
    with open('ANX_ITEMS.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['item = nr zlecenia','material','data rozp.','waga','grupa mat.'])
        write.writerows(rows)

    if subset == 'J2':
        return items, indJ2, indJ2_1, indJ2_2
    elif subset == 'J3':
        return items, indJ3

def edit_open_orders(sales,group,prefLines,prods):
    items1 = []
    items2 = []

    rows1 = []
    rows2 = []
    
    indJ0 = []
    indJ1 = []
    
    minDDS = 999999
    
    for row in sales:
        nbr = str(row['Zlecenie'])
        if nbr in ['None','']:
            continue

        material = str(row['Indeks'])
        matName = str(row['Nazwa'])

        someday = row['Plan'].date()
        date_r = str(someday)
        
        dds = float((someday - TODAY).days)
        if minDDS > dds:
            minDDS = dds
        try:
            weight = float(row['Suma z do zrealizowania1'])
        except:
            print('WARNING: błędne dane albo brak danych o wykonaniu dla zlecenia (brak wagi?)', nbr)
            weight = float(-1)
        grupa = group.get(material,'brak')
        podz = 0
        prior = float(dds)

        try:
            weight_done = float(row['Suma z Ilość zreal.[KG]'])
        except:
            print('WARNING: błędne dane albo brak danych o wykonaniu dla zlecenia (brag wagi wykonanej?)', nbr)
            weight_done=float(-1)
        
        if material in prods: #alert komentarz z 02.11, materiały spoza listy w ogóle pomijamy
            if weight > 0: # w ogóle jest coś do roboty
                if weight_done > 0: # coś już zrobiono - zlecenie otwarte rozpoczęte
                    try:
                        nPrefLine = prefLines[material]
                    except:
                        nPrefLine = 'brak'
                        print('WARNING: Niezdefiniowana linia preferowana dla indx ',material, 'przyjmuję brak')
                    finally:
                        if material in matIndxMix:
                            wgt1 = weight * MAT_MIX[matIndxMix.index(material)].udzial1
                            wgt2 = weight * MAT_MIX[matIndxMix.index(material)].udzial2
                            lpi1 =  MAT_MIX[matIndxMix.index(material)].lPrio1
                            lpi2 =  MAT_MIX[matIndxMix.index(material)].lPrio2
                            items1.append(ZLECENIE(nbr+'m1',material,date_r,wgt1, wgt1,grupa,podz,'J0',lpi1,prior,matName))
                            items1.append(ZLECENIE(nbr+'m2',material,date_r,wgt2, wgt2,grupa,podz,'J0',lpi2,prior,matName))
                            indJ0.append(nbr+'m1')
                            indJ0.append(nbr+'m2')
                            item = items1[-2]
                            rows1.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset,item.prefLine,item.priority])
                            item = items1[-1]
                            rows1.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset,item.prefLine,item.priority])
                        else:
                            items1.append(ZLECENIE(nbr,material,date_r,weight, weight,grupa,podz,'J0',nPrefLine,prior,matName))
                            indJ0.append(nbr)
                            item = items1[-1]
                            rows1.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset,item.prefLine,item.priority])
                else: # zlecenie otwarte nierozpoczęte
                    try:
                        nPrefLine = prefLines[material]
                    except:
                        nPrefLine = 'brak'
                        print('WARNING: Niezdefiniowana linia preferowana dla indx ',material, 'przyjmuję brak')
                    finally:
                        if material in matIndxMix:
                            wgt1 = weight * MAT_MIX[matIndxMix.index(material)].udzial1
                            wgt2 = weight * MAT_MIX[matIndxMix.index(material)].udzial2
                            lpi1 =  MAT_MIX[matIndxMix.index(material)].lPrio1
                            lpi2 =  MAT_MIX[matIndxMix.index(material)].lPrio2
                            items2.append(ZLECENIE(nbr+'m1',material,date_r,wgt1,wgt1,grupa,podz,'J1',lpi1,prior,matName))
                            items2.append(ZLECENIE(nbr+'m2',material,date_r,wgt2,wgt2,grupa,podz,'J1',lpi2,prior,matName))
                            indJ1.append(nbr+'m1')
                            indJ1.append(nbr+'m2')
                            item = items2[-2]
                            rows2.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset,item.prefLine,item.priority])
                            item = items2[-1]
                            rows2.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset,item.prefLine,item.priority])
                        else:
                            items2.append(ZLECENIE(nbr,material,date_r,weight,weight,grupa,podz,'J1',nPrefLine,prior,matName))
                            indJ1.append(nbr)
                            item = items2[-1]
                            rows2.append([item.nbr, item.material, item.date_r, item.weight,item.originalWeight,item.group,item.divisible,item.indSubset,item.prefLine,item.priority])
    if minDDS < 1:
        for its in range(0,len(items1)):
            items1[its].priority = float(1) / (float(items1[its].priority) - float(minDDS) + float(1))
        for its in range(0,len(items2)):
            items2[its].priority = float(1) / (float(items2[its].priority) - float(minDDS) + float(1))
    else:
        for its in range(0,len(items1)):
            items1[its].priority = float(1) / float(items1[its].priority)
        for its in range(0,len(items2)):
            items2[its].priority = float(1) / float(items2[its].priority)

    with open('ANX_ITEMS1.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['item = nr zlecenia','material','data rozp.','waga','grupa mat.','podz','subset','prefLine','prio'])
        write.writerows(rows1)
        
    with open('ANX_ITEMS2.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['item = nr zlecenia','material','data rozp.','waga','grupa mat.','podz','subset','prefLine','prio'])
        write.writerows(rows2)

    return items1, items2, indJ0, indJ1

def edit_mat_mix(mixProds):
    items = []
    rows = []
    mixIndexList = []

    for row in mixProds:
        mat = str(row['Materiał'])
        if mat in ['None','']:
            continue

        lPrior1 = row['linia prio 1']
        lPrior2 = row['linia prio 2']
        udzi1 = row['udzial 1']
        udzi2 = row['udzial 2']
        
        items.append(MIX_PROD(mat, lPrior1, lPrior2, udzi1, udzi2))
        mixIndexList.append(mat)
        item = items[-1]
        rows.append([item.material, item.lPrio1, item.lPrio2, item.udzial1, item.udzial2])

    with open('ANX_MIXES.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['material','linia prio 1','linia prio 2','udział 1','udział 2'])
        write.writerows(rows)

    return items, mixIndexList

def findMatProd(itemsList,indxList):
    matProdList = []

    for j in range(len(indxList)):
        mat = itemsList[j].material
        if mat not in matProdList:
            matProdList.append(mat)
    
    return matProdList

def check_production_version(items,products):
    braki = []
    for item in items:
        if item.material not in products:
            braki.append([item.nbr, item.material])

    with open('ANX_BRAKI_PW.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['nr zlecenia','material'])
        write.writerows(braki)

    return braki

def check_MP(items,products,mp,eff,lines):
    braki = []
    for item1 in items:
        if item1.material not in products:
            continue
        if item1.group in [None,'None','','brak']:
            braki.append([ (item1.nbr,item1.material),'brak przypisanej grupy produktowej'])
        for item2 in items:
            if item2.material == item1.material:
                continue
            for l in lines:
                if (eff.get((item1.material,l),None) not in [None,0]) and (eff.get((item2.material,l),None) not in [None,0]):
                    if mp.get((item1.material,item2.material,l),None) == None:
                        braki.append([(item1.material,item2.material,l),'brak przezbrojenia w tabelach'])
                    if mp.get((item1.material, item2.material, l), None) == 0:
                        braki.append([(item1.material, item2.material, l), 'przezbrojenie = 0'])

    with open('ANX_BRAKI_MP.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['obiekt','uwagi'])
        write.writerows(braki)

    return braki

def edit_mat_spec(material_specification):
    components = {}
    consumption = {}

    for row in material_specification:
        mat = row['Materiał']
        skl = row['Składnik']
        zuzycie = float(row['Il. na 1000KG'])/1000

        try:
            components[mat].append(skl)
        except KeyError:
            components[mat] = [skl]

        consumption[(mat,skl)] = zuzycie


    rows = []
    for cmp in components:
        rows.append([cmp,str(components[cmp])])
    with open('ANX_COMPONENTS.csv', 'w') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['material','skladniki'])
        write.writerows(rows)

    rows = []
    for cns in consumption:
        rows.append([str(cns), consumption[cns]])
    with open('ANX_CONSUMPTION.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['(material,skladnik)','zuzycie'])
        write.writerows(rows)

    return components, consumption

def edit_stock(stock_levels):
    stock = {}

    for row in stock_levels:
        mat = row['Materiał']
        il = float(row['Dowol. wyk. wart.'])

        stock[mat] = stock.get(mat,0) + il

    return stock

def edit_tps(czasy_tps):
    tps = {}
    for row in czasy_tps:
        tps[str(row['Materiał'])] = int(row['dni na wyrotowanie'])

    return tps

def check_tps(items,tps):
    braki = []

    for item in items:
        if item.material not in tps:
            braki.append([item.nbr, item.material])

    with open('ANX_BRAKI_TPS.csv', 'w', encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['nr zlecenia', 'material'])
        write.writerows(braki)

    return braki

PATHS_TO_DATA = loadParameters(pathParam, 'paths',1)

PARAM_GLOBAL = loadParameters(pathParam, 'global',1)

minQij = loadParameters(pathParam, 'qij',2)

shiftsData, indxSH = loadParameters(pathParam, 'shifts',3)

def computeShiftStartTimes(sData):
    shTimes=[]
    sTm = 0.0
    shTimes.append(sTm)
    for sh in range(0,len(indxSH)-1):
        sTm += shiftsData[str(indxSH[sh]),'maxTotalTime']
        shTimes.append(sTm)
    return shTimes

tShiftStart = computeShiftStartTimes(shiftsData)

linesData, secondIndxLines = loadParameters(pathParam, 'lines',8)

gC1 = loadParameters(pathParam, 'changeover',1)
groupChangeData, secondIndxGroup = loadParameters(pathParam, 'changeover',len(gC1)+1)

def getNonZeros(lData, sKey,xcolumn):
    LL = []
    for key in sKey:
        if lData[key,xcolumn] != 0:
            LL.append(key)
    return LL

indxLimitedLINES = getNonZeros(linesData, secondIndxLines, 'do_limitu')

indxSHIFTS = getNonZeros(shiftsData,indxSH,'nShifts')

ZL_PROD = load_data_from_excel(PATHS_TO_DATA['zlecenia produkcyjne'], 'Właściwy1',1,1)

ZL_PLAN = load_data_from_excel(PATHS_TO_DATA['zlecenia planowane'], 'Właściwy1',1,1)

ZL_OTW = load_data_from_excel(PATHS_TO_DATA['zlecenia otwarte'], 'pakowanie',4,1)

MT_MIX = load_data_from_excel(PATHS_TO_DATA['plastrownia'],'MIX',3,1)

MAT_MIX, matIndxMix = edit_mat_mix(MT_MIX)

EFFICIENCY, indxLINES, GROUP, PRODUCTS, PREFERENCES, prodTime = load_efficiency(PATHS_TO_DATA['plastrownia'],'wydajności')

TPS_TIMES = load_data_from_excel(PATHS_TO_DATA['tps'], 'Arkusz1',1,1)

TPS = edit_tps(TPS_TIMES)

ITEMS_PROD, indxJ2, indxJ2_1, indxJ2_2 = edit_sales_orders(ZL_PROD, GROUP, 'J2', PREFERENCES,0,0,TPS)

ITEMS_PLAN, indxJ3 = edit_sales_orders(ZL_PLAN, GROUP, 'J3', PREFERENCES,PARAM_GLOBAL['waga_termin'],PARAM_GLOBAL['waga_ilosc'],TPS)

indxJdummy = ['dummy']

ITEMS_OTW_ROZP,ITEMS_OTW_NIEROZP, indxJ0, indxJ1 = edit_open_orders(ZL_OTW, GROUP, PREFERENCES, PRODUCTS)

indxJ = indxJ0 + indxJ1 + indxJ2
indxJExt = indxJdummy + indxJ
indxJLong = indxJ + indxJ3
indxJExtLong = indxJExt + indxJ3

ITEMS_ALL = ITEMS_OTW_ROZP + ITEMS_OTW_NIEROZP + ITEMS_PROD + ITEMS_PLAN

MAT_PROD = findMatProd(ITEMS_ALL,indxJ)

MP = load_MP(PATHS_TO_DATA['plastrownia'],GROUP,indxLINES,groupChangeData)

#---------------podłączenie do serwera-----------------------
with gp.Env(empty=True) as env:
    env.setParam('CSManager', 'http://10.101.1.151:61080')
    env.setParam('CSAPIACCESSID', '35aa618d-b8e1-48d7-81c9-588b771991bd')
    env.setParam('CSAPISECRET', '6fd7cb7d-567f-47ad-9868-2f88d4a7fd10')
    env.setParam('CSAPPNAME', 'packApp')
    env.start()
    with gp.Model(env=env,name='ANX_Daily') as myModel:

    #myModel = gp.Model('ANX_Daily')

        w_i = len(indxLINES)*['']
        v_i = len(indxLINES)*['']
        n_i = len(indxLINES)*['']
        n_i0 = len(indxLINES)*['']
        T1_i = len(indxLINES)*['']
        T2_i = len(indxLINES)*['']
        tStart_i = len(indxLINES)*['']
        tStop_i = len(indxLINES)*['']
        u_is = [['' for x in range(len(indxSHIFTS))] for y in range (len(indxLINES))]
        uPrime_is = [['' for x in range(len(indxSHIFTS))] for y in range (len(indxLINES))]
        x_ij = [['' for x in range(len(indxJLong))] for y in range (len(indxLINES))]
        t_ij = [['' for x in range(len(indxJLong))] for y in range (len(indxLINES))]
        z_ij = [['' for x in range(len(indxJLong))] for y in range (len(indxLINES))]
        p_ij = [['' for x in range(len(indxJLong))] for y in range (len(indxLINES))]
        y_ijk = [[['' for x in range(-1,len(indxJLong))] for y in range(-1,len(indxJLong))] for z in range (len(indxLINES))]
        q_j = len(indxJLong)*['']
        b_ij = [['' for x in range(len(indxJLong))] for y in range (len(indxLINES))]
        a_j = len(indxJLong)*['']

        for i in range(len(indxLINES)):
            w_i[i] = myModel.addVar(vtype=GRB.INTEGER, name="w_"+indxLINES[i])
            v_i[i] = myModel.addVar(vtype=GRB.BINARY, name="v_"+indxLINES[i])
            n_i[i] = myModel.addVar(vtype=GRB.INTEGER, name="n_"+indxLINES[i])
            n_i0[i] = myModel.addVar(vtype=GRB.INTEGER, name="n_"+indxLINES[i])
            T1_i[i] = myModel.addVar(vtype=GRB.CONTINUOUS, name="T1_"+indxLINES[i])
            T2_i[i] = myModel.addVar(vtype=GRB.CONTINUOUS, name="T2_"+indxLINES[i])
            tStart_i[i] = myModel.addVar(vtype=GRB.CONTINUOUS, name="tStart_"+indxLINES[i])
            tStop_i[i] = myModel.addVar(vtype=GRB.CONTINUOUS, name="tStop_"+indxLINES[i])
            for s in range(len(indxSHIFTS)):
                u_is[i][s] = myModel.addVar(vtype=GRB.BINARY, name="u_"+indxLINES[i]+','+indxSHIFTS[s])
                uPrime_is[i][s] = myModel.addVar(vtype=GRB.BINARY, name="uPrime_"+indxLINES[i]+','+indxSHIFTS[s])
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    x_ij[i][j] = myModel.addVar(vtype=GRB.CONTINUOUS, name="x_"+indxLINES[i]+','+indxJLong[j])
                    t_ij[i][j] = myModel.addVar(vtype=GRB.CONTINUOUS, name="t_"+indxLINES[i]+','+indxJLong[j])
                    z_ij[i][j] = myModel.addVar(vtype=GRB.BINARY, name="z_"+indxLINES[i]+','+indxJLong[j])
                    p_ij[i][j] = myModel.addVar(vtype=GRB.CONTINUOUS, name="p_"+indxLINES[i]+','+indxJLong[j])
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    y_ijk[i][j][-1] = myModel.addVar(vtype=GRB.BINARY, name="y_"+indxLINES[i]+','+indxJLong[j]+','+indxJdummy[0])
                    y_ijk[i][-1][j] = myModel.addVar(vtype=GRB.BINARY, name="y_"+indxLINES[i]+','+indxJdummy[0]+','+indxJLong[j])
                    for k in range(len(indxJLong)):
                        if (ITEMS_ALL[k].material,indxLINES[i]) in EFFICIENCY:
                            y_ijk[i][j][k] = myModel.addVar(vtype=GRB.BINARY, name="y_"+indxLINES[i]+','+indxJLong[j]+','+indxJLong[k])
        for j in range(len(indxJLong)):
            q_j[j] = myModel.addVar(vtype=GRB.BINARY,name="q_"+indxJLong[j])
            
            
        for j in range(len(indxJLong)):
            if ITEMS_ALL[j].material in matIndxMix:
                if ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm1':
                    if ITEMS_ALL[j+1].nbr[-2:len(ITEMS_ALL[j+1].nbr)] == 'm2': # podwójne zabezpieczenie
                        if ITEMS_ALL[j+1].nbr[0:len(ITEMS_ALL[j+1].nbr)-2] == ITEMS_ALL[j].nbr[0:len(ITEMS_ALL[j].nbr)-2]: # a nawet potrójne
                            a_j[j] = myModel.addVar(vtype=GRB.BINARY,name="b_"+indxJLong[j])
                            for i in range(len(indxLINES)):
                                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                                    if (ITEMS_ALL[j+1].material,indxLINES[i]) in EFFICIENCY:
                                        b_ij[i][j] = myModel.addVar(vtype=GRB.BINARY, name="a_"+indxLINES[i]+','+indxJLong[j])

        exprObj0=LinExpr()
        for i in range(len(indxLINES)):
            for j in range(len(indxJ0)):
                jj = indxJLong.index(indxJ0[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprObj0.addTerms(PARAM_GLOBAL['bigM']**4,x_ij[i][jj])
            for j in range(len(indxJ1)):
                jj = indxJLong.index(indxJ1[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprObj0.addTerms(PARAM_GLOBAL['bigM']**4,x_ij[i][jj])       
        exprObj1=LinExpr()
        for i in range(len(indxLINES)):
            for j in range(len(indxJ2)):
                jj = indxJLong.index(indxJ2[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprObj1.addTerms((PARAM_GLOBAL['bigM'])**3,x_ij[i][jj])
        exprObj2=LinExpr()
        for i in range(len(indxLINES)):
            for j in range(len(indxJ3)):
                jj = indxJLong.index(indxJ3[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprObj2.addTerms(ITEMS_PLAN[j].priority,x_ij[i][jj])
                    
        exprObj3=LinExpr()
        for i in range(len(indxLINES)):
            for j in range(len(indxJ0)):
                if indxLINES[i] in ITEMS_OTW_ROZP[j].prefLine.split(';'): #alert AZ
                    jj = indxJLong.index(indxJ0[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])
            for j in range(len(indxJ1)):
                if indxLINES[i] in ITEMS_OTW_NIEROZP[j].prefLine.split(';'):  # alert AZ
                    jj = indxJLong.index(indxJ1[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])
            for j in range(len(indxJ2)):
                if indxLINES[i] in ITEMS_PROD[j].prefLine.split(';'):  # alert AZ
                    jj = indxJLong.index(indxJ2[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms((PARAM_GLOBAL['bigM'])**3,x_ij[i][jj])
            for j in range(len(indxJ3)):
                if indxLINES[i] in ITEMS_PLAN[j].prefLine.split(';'):  # alert AZ
                    jj = indxJLong.index(indxJ3[j])
                    if ITEMS_ALL[jj].material not in matIndxMix:
                        if ITEMS_ALL[jj].material in MAT_PROD:
                            if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                                exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])            
                    
        """            
        exprObj3=LinExpr()
        for i in range(len(indxLINES)):
            for j in range(len(indxJ0)):
                if ITEMS_OTW_ROZP[j].prefLine == indxLINES[i]:
                    jj = indxJLong.index(indxJ0[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])
            for j in range(len(indxJ1)):
                if ITEMS_OTW_NIEROZP[j].prefLine == indxLINES[i]:
                    jj = indxJLong.index(indxJ1[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])
            for j in range(len(indxJ2)):
                if ITEMS_PROD[j].prefLine == indxLINES[i]:
                    jj = indxJLong.index(indxJ2[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms((PARAM_GLOBAL['bigM'])**3,x_ij[i][jj])
            for j in range(len(indxJ3)):
                if ITEMS_PLAN[j].prefLine == indxLINES[i]:
                    jj = indxJLong.index(indxJ3[j])
                    if ITEMS_ALL[jj].material not in matIndxMix:
                        if ITEMS_ALL[jj].material in MAT_PROD:
                            if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                                exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])"""
                        
        exprObj4=LinExpr()                
        for j in range(len(indxJLong)):
            if ITEMS_ALL[j].material in matIndxMix:
                if ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm1':
                    if ITEMS_ALL[j+1].nbr[-2:len(ITEMS_ALL[j+1].nbr)] == 'm2': # podwójne zabezpieczenie
                        if ITEMS_ALL[j+1].nbr[0:len(ITEMS_ALL[j+1].nbr)-2] == ITEMS_ALL[j].nbr[0:len(ITEMS_ALL[j].nbr)-2]: # a nawet potrójne
                            exprObj4.addTerms(-((PARAM_GLOBAL['bigM'])**3)*(ITEMS_ALL[j].weight+ITEMS_ALL[j+1].weight),a_j[j])

        exprObj5=LinExpr()
        for i in range(len(indxLINES)):
            exprObj5.addTerms(-PARAM_GLOBAL['bigMPenalty'],T1_i[i])
            
        exprObj6=LinExpr()
        for i in range(len(indxLINES)):
            for j in range(len(indxJ0)):
                if ITEMS_OTW_ROZP[j].prefLine == 'brak':
                    jj = indxJLong.index(indxJ0[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])
            for j in range(len(indxJ1)):
                if ITEMS_OTW_NIEROZP[j].prefLine == 'brak':
                    jj = indxJLong.index(indxJ1[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])
            for j in range(len(indxJ2)):
                if ITEMS_PROD[j].prefLine == 'brak':
                    jj = indxJLong.index(indxJ2[j])
                    if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                        exprObj3.addTerms((PARAM_GLOBAL['bigM'])**3,x_ij[i][jj])
            for j in range(len(indxJ3)):
                if ITEMS_PLAN[j].prefLine == 'brak':
                    jj = indxJLong.index(indxJ3[j])
                    if ITEMS_ALL[jj].material not in matIndxMix:
                        if ITEMS_ALL[jj].material in MAT_PROD:
                            if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                                exprObj3.addTerms(PARAM_GLOBAL['bigM'],x_ij[i][jj])
                            
        exprObj=LinExpr()
        exprObj = exprObj0 + exprObj1 + exprObj2 + exprObj3 + exprObj4 + exprObj5 + exprObj6
        myModel.setObjective(exprObj, GRB.MAXIMIZE)  

        exprTi1=len(indxLINES)*['']
        for i in range(len(indxLINES)):
            exprTi1[i]=LinExpr()
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    for k in range(len(indxJLong)):
                        if (ITEMS_ALL[k].material,indxLINES[i]) in EFFICIENCY:
                            try:
                                changeover_ijk = MP[(ITEMS_ALL[j].material,ITEMS_ALL[k].material,indxLINES[i])]
                            except:
                                print('WARNING: brak czasu przezbrojenia',',',j,',',k,',',i,',',ITEMS_ALL[j].material,',',ITEMS_ALL[k].material,',',indxLINES[i],',',changeover_ijk)
                                changeover_ijk = 0
                            finally:
                                exprTi1[i].addTerms(changeover_ijk,y_ijk[i][j][k])
            myModel.addConstr(T1_i[i] == exprTi1[i])

        exprTi2=len(indxLINES)*['']
        for i in range(len(indxLINES)):
            exprTi2[i]=LinExpr()
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    try:
                        prodtime_ij = prodTime[(ITEMS_ALL[j].material,indxLINES[i])]
                    except:
                        print('WARNING: brak czasu produkcji',',',j,',',i,',',ITEMS_ALL[j].material,',',indxLINES[i],',',prodtime_ij)
                        prodtime_ij = PARAM_GLOBAL['bigM']
                    finally:
                        exprTi2[i].addTerms(prodtime_ij,x_ij[i][j])
            myModel.addConstr(T2_i[i] == exprTi2[i])

        exprMaxTTi=len(indxLINES)*['']
        for i in range(len(indxLINES)):
            exprMaxTTi[i]=LinExpr()
            for s in range(len(indxSHIFTS)):
                try:
                    maxTTs = shiftsData[(indxSHIFTS[s],'maxTotalTime')]
                except:
                    print('WARNING: brak czasu maksymalnego dla linii',',',s,',',i)
                    maxTTs = 0
                finally:
                    exprMaxTTi[i].addTerms(maxTTs,u_is[i][s])
            myModel.addConstr(exprTi1[i] + exprTi2[i] <= exprMaxTTi[i])

        for i in range(len(indxLINES)):
            myModel.addConstr(v_i[i] <= w_i[i])
            myModel.addConstr(w_i[i] <= len(indxSHIFTS) * v_i[i])

        exprUIS = len(indxLINES)*[''] 
        for i in range(len(indxLINES)):
            exprUIS[i] = LinExpr()
            for s in range(len(indxSHIFTS)):
                exprUIS[i].addTerms(1,u_is[i][s])
            myModel.addConstr(exprUIS[i] == w_i[i])

        for i in range(len(indxLINES)):
            for s in range(len(indxSHIFTS)):
                myModel.addConstr(u_is[i][s] <= linesData[indxLINES[i],indxSHIFTS[s]])

        for i in range(len(indxLINES)):
            for r in range(len(indxSHIFTS)-2):
                if linesData[indxLINES[i],indxSHIFTS[r]] == 1:
                    for s in range(r+1,len(indxSHIFTS)-1):
                        if linesData[indxLINES[i],indxSHIFTS[s]] == 1:
                            for t in range(s+1,len(indxSHIFTS)):
                                if linesData[indxLINES[i],indxSHIFTS[t]] == 1:
                                    myModel.addConstr(u_is[i][r] - u_is[i][s] + u_is[i][t] <= 1)

        exprUi0s = len(indxSHIFTS)*[''] 
        for s in range(len(indxSHIFTS)):
            exprUi0s[s] = LinExpr()
            for i in range(len(indxLimitedLINES)):
                ii = indxLINES.index(indxLimitedLINES[i])
                exprUi0s[s].addTerms(1,u_is[ii][s])
            myModel.addConstr(exprUi0s[s] == shiftsData[(indxSHIFTS[s],'nShifts')])

        exprZij = len(indxJLong)*['']
        for j in range(len(indxJLong)):
            exprZij[j] = LinExpr()
            for i in range(len(indxLINES)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    exprZij[j].addTerms(1,z_ij[i][j])
            myModel.addConstr(exprZij[j] <= len(indxLINES)*q_j[j])
            myModel.addConstr(exprZij[j] >= q_j[j])

        for i in range(len(indxLINES)):
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    if (indxLINES[i],ITEMS_ALL[j].material) in minQij:
                        mQij = minQij[indxLINES[i],ITEMS_ALL[j].material]
                    elif ('default',ITEMS_ALL[j].material) in minQij:
                        mQij = minQij['default',ITEMS_ALL[j].material]
                    elif (indxLINES[i],'default') in minQij:
                        mQij = minQij[indxLINES[i],'default']
                    elif ('default','default') in minQij:
                        mQij = minQij['default','default']
                    else:
                        mQij = 500
                    #modyfikacja proporcjonalnie do udziałów, gdy MIX
                    if ITEMS_ALL[j].material in matIndxMix:
                        if ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm1':
                            mQij *= MAT_MIX[matIndxMix.index(ITEMS_ALL[j].material)].udzial1
                        elif ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm2':
                            mQij *= MAT_MIX[matIndxMix.index(ITEMS_ALL[j].material)].udzial2
                    #jeżeli zlecenie zaległe to modyfikacja qij
                    if ITEMS_ALL[j].weight < mQij:
                        if ITEMS_ALL[j].indSubset in ['J0','J1']:
                            mQij = ITEMS_ALL[j].weight
                        elif ITEMS_ALL[j].indSubset in ['J2_1','J2_2','J3']:
                            ITEMS_ALL[j].weight = mQij
                    myModel.addConstr(mQij * z_ij[i][j] <= x_ij[i][j])
                    myModel.addConstr(x_ij[i][j] <= ITEMS_ALL[j].weight * z_ij[i][j])

        exprJ0Xij = len(indxJ0)*[''] 
        for j in range(len(indxJ0)):
            exprJ0Xij[j] = LinExpr()
            jj = indxJLong.index(indxJ0[j])
            for i in range(len(indxLINES)):
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprJ0Xij[j].addTerms(1,x_ij[i][jj])
            myModel.addConstr(exprJ0Xij[j] <= ITEMS_ALL[jj].weight * q_j[jj]) # alert relaksacja z 16.11
            
        exprJ1Xij = len(indxJ1)*[''] 
        for j in range(len(indxJ1)):
            exprJ1Xij[j] = LinExpr()
            jj = indxJLong.index(indxJ1[j])
            for i in range(len(indxLINES)):
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprJ1Xij[j].addTerms(1,x_ij[i][jj])
            myModel.addConstr(exprJ1Xij[j] <= ITEMS_ALL[jj].weight * q_j[jj]) # alert relaksacja z 16.11

        exprJ2_1Xij = len(indxJ2_1)*[''] 
        for j in range(len(indxJ2_1)):
            exprJ2_1Xij[j] = LinExpr()
            jj = indxJLong.index(indxJ2_1[j])
            for i in range(len(indxLINES)):
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprJ2_1Xij[j].addTerms(1,x_ij[i][jj])
            myModel.addConstr(exprJ2_1Xij[j] == ITEMS_ALL[jj].weight) # alert uwagi z 15.12: muszą być w 100% wykonane danego dnia
            # myModel.addConstr(exprJ2_1Xij[j] == ITEMS_ALL[jj].weight * q_j[jj]) # alert stara wersja - działa, gdy przesadzono z ograniczeniami

        exprJ2_2Xij = len(indxJ2_2)*[''] 
        for j in range(len(indxJ2_2)):
            exprJ2_2Xij[j] = LinExpr()
            jj = indxJLong.index(indxJ2_2[j])
            for i in range(len(indxLINES)):
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprJ2_2Xij[j].addTerms(1,x_ij[i][jj])
            myModel.addConstr(exprJ2_2Xij[j] <= ITEMS_ALL[jj].weight * q_j[jj])
            
        exprJ3Xij = len(indxJ3)*[''] 
        for j in range(len(indxJ3)):
            exprJ3Xij[j] = LinExpr()
            jj = indxJLong.index(indxJ3[j])
            for i in range(len(indxLINES)):
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprJ3Xij[j].addTerms(1,x_ij[i][jj])
            myModel.addConstr(exprJ3Xij[j] <= ITEMS_ALL[jj].weight * q_j[jj])

        exprYijk = [['' for x in range(len(indxJLong))] for y in range (len(indxLINES))]
        exprYikj = [['' for x in range(len(indxJLong))] for y in range (len(indxLINES))]
        for i in range(len(indxLINES)):
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    exprYijk[i][j] = LinExpr()
                    exprYikj[i][j] = LinExpr()
                    for k in range(len(indxJLong)):
                        if (ITEMS_ALL[k].material,indxLINES[i]) in EFFICIENCY:
                            exprYijk[i][j].addTerms(1,y_ijk[i][j][k])
                            exprYikj[i][j].addTerms(1,y_ijk[i][k][j])
                    exprYijk[i][j].addTerms(1,y_ijk[i][j][-1])
                    exprYikj[i][j].addTerms(1,y_ijk[i][-1][j])
                    myModel.addConstr(exprYijk[i][j] == z_ij[i][j])
                    myModel.addConstr(exprYikj[i][j] == z_ij[i][j])

        exprYi0k = len(indxJLong)*['']
        exprYik0 = len(indxJLong)*['']
        for i in range(len(indxLINES)):
            exprYi0k[i] = LinExpr()
            exprYik0[i] = LinExpr()
            for k in range(len(indxJLong)):
                if (ITEMS_ALL[k].material,indxLINES[i]) in EFFICIENCY:
                    exprYi0k[i].addTerms(1,y_ijk[i][-1][k])
                    exprYik0[i].addTerms(1,y_ijk[i][k][-1])
            myModel.addConstr(exprYi0k[i] == v_i[i])
            myModel.addConstr(exprYik0[i] == v_i[i])

        exprZijNi = len(indxLINES) * ['']
        exprZijNi0 = len(indxLINES) * [''] #dodane, żeby linia się  nie włączała dla samych planowanych
        for i in range(len(indxLINES)):
            exprZijNi[i] = LinExpr()
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    exprZijNi[i].addTerms(1,z_ij[i][j])
            myModel.addConstr(exprZijNi[i] == n_i[i])
            #dodane, żeby linia się  nie włączała dla samych planowanych:
            exprZijNi0[i] = LinExpr()
            for j in range(len(indxJ)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    exprZijNi0[i].addTerms(1,z_ij[i][j])
            myModel.addConstr(exprZijNi0[i] == n_i0[i])
            myModel.addConstr(n_i0[i] >= v_i[i]) #wariant 1
            #alternatywnie
            #myModel.addConstr(n_i0[i]*len(indxJLong) >= n_i[i]) #wariant 2

        for i in range(len(indxLINES)):
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    myModel.addConstr(z_ij[i][j] <= p_ij[i][j])
                    myModel.addConstr(p_ij[i][j] <= len(indxJLong) * z_ij[i][j])
                    myModel.addConstr(p_ij[i][j] <= n_i[i])

        for i in range(len(indxLINES)):
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                     for k in range(len(indxJLong)):
                        if (ITEMS_ALL[k].material,indxLINES[i]) in EFFICIENCY:
                            myModel.addConstr(p_ij[i][j] - p_ij[i][k] + len(indxJLong) * y_ijk[i][j][k] <= len(indxJLong) - z_ij[i][k])

        for i in range(len(indxLINES)):
            for j in range(len(indxJ0)):
                jj = indxJLong.index(indxJ0[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    for k in range(len(indxJ1)):
                        kk = indxJLong.index(indxJ1[k])
                        if (ITEMS_ALL[kk].material,indxLINES[i]) in EFFICIENCY:
                            if ITEMS_ALL[jj].group != 'WA' or ITEMS_ALL[kk].group == 'WA':
                                if ITEMS_ALL[jj].group == 'DR1' or ITEMS_ALL[kk].group != 'DR1':
                                    if ITEMS_ALL[jj].group == 'WP1' or ITEMS_ALL[kk].group != 'WP1':
                                        myModel.addConstr(p_ij[i][jj] <= p_ij[i][kk] + len(indxJLong) * (1 - z_ij[i][kk]))
                
        for i in range(len(indxLINES)):
            for j in range(len(indxJ0)):
                jj = indxJLong.index(indxJ0[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    for k in range(len(indxJ2)):
                        kk = indxJLong.index(indxJ2[k])
                        if (ITEMS_ALL[kk].material,indxLINES[i]) in EFFICIENCY:
                            if ITEMS_ALL[jj].group != 'WA' or ITEMS_ALL[kk].group == 'WA':
                                if ITEMS_ALL[jj].group == 'DR1' or ITEMS_ALL[kk].group != 'DR1':
                                    if ITEMS_ALL[jj].group == 'WP1' or ITEMS_ALL[kk].group != 'WP1':
                                        myModel.addConstr(p_ij[i][jj] <= p_ij[i][kk] + len(indxJLong) * (1 - z_ij[i][kk]))
                            
        for i in range(len(indxLINES)):
            for j in range(len(indxJ0)):
                jj = indxJLong.index(indxJ0[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    for k in range(len(indxJ3)):
                        kk = indxJLong.index(indxJ3[k])
                        if (ITEMS_ALL[kk].material,indxLINES[i]) in EFFICIENCY:
                            if ITEMS_ALL[jj].group != 'WA' or ITEMS_ALL[kk].group == 'WA':
                                if ITEMS_ALL[jj].group == 'DR1' or ITEMS_ALL[kk].group != 'DR1':
                                    if ITEMS_ALL[jj].group == 'WP1' or ITEMS_ALL[kk].group != 'WP1':
                                        myModel.addConstr(p_ij[i][jj] <= p_ij[i][kk] + len(indxJLong) * (1 - z_ij[i][kk]))
                            
        for i in range(len(indxLINES)):
            for j in range(len(indxJ1)):
                jj = indxJLong.index(indxJ1[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    for k in range(len(indxJ2)):
                        kk = indxJLong.index(indxJ2[k])
                        if (ITEMS_ALL[kk].material,indxLINES[i]) in EFFICIENCY:
                            if ITEMS_ALL[jj].group != 'WA' or ITEMS_ALL[kk].group == 'WA':
                                if ITEMS_ALL[jj].group == 'DR1' or ITEMS_ALL[kk].group != 'DR1':
                                    if ITEMS_ALL[jj].group == 'WP1' or ITEMS_ALL[kk].group != 'WP1':
                                        myModel.addConstr(p_ij[i][jj] <= p_ij[i][kk] + len(indxJLong) * (1 - z_ij[i][kk]))
                            
        for i in range(len(indxLINES)):
            for j in range(len(indxJ1)):
                jj = indxJLong.index(indxJ1[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    for k in range(len(indxJ3)):
                        kk = indxJLong.index(indxJ3[k])
                        if (ITEMS_ALL[kk].material,indxLINES[i]) in EFFICIENCY:
                            if ITEMS_ALL[jj].group != 'WA' or ITEMS_ALL[kk].group == 'WA':
                                if ITEMS_ALL[jj].group == 'DR1' or ITEMS_ALL[kk].group != 'DR1':
                                    if ITEMS_ALL[jj].group == 'WP1' or ITEMS_ALL[kk].group != 'WP1':
                                        myModel.addConstr(p_ij[i][jj] <= p_ij[i][kk] + len(indxJLong) * (1 - z_ij[i][kk]))

        # Chwilowo wyłączone, kolejność między produkcyjnymi a planowanymi. Stara wersja, więc nie uwzględnia DR1 ani WP1
        # for i in range(len(indxLINES)):
        #     for j in range(len(indxJ2)):
        #         jj = indxJLong.index(indxJ2[j])
        #         if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
        #             for k in range(len(indxJ3)):
        #                 kk = indxJLong.index(indxJ3[k])
        #                 if (ITEMS_ALL[kk].material,indxLINES[i]) in EFFICIENCY:
        #                     if ITEMS_ALL[jj].group != 'WA' or ITEMS_ALL[kk].group == 'WA':
        #                         myModel.addConstr(p_ij[i][jj] <= p_ij[i][kk] + len(indxJLong) * (1 - z_ij[i][kk]))

        exprZijJ0J1 = len(indxLINES) * ['']
        for i in range(len(indxLINES)):
            exprZijJ0J1[i] = LinExpr()
            for j in range(len(indxJ0)):
                jj = indxJLong.index(indxJ0[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprZijJ0J1[i].addTerms(1,z_ij[i][jj])
            for j in range(len(indxJ1)):
                jj = indxJLong.index(indxJ1[j])
                if (ITEMS_ALL[jj].material,indxLINES[i]) in EFFICIENCY:
                    exprZijJ0J1[i].addTerms(1,z_ij[i][jj])
            myModel.addConstr(exprZijJ0J1[i] <= (len(indxJ0) + len(indxJ1)) * u_is[i][0])

        exprUPrimeIS = len(indxLINES)*[''] 
        for i in range(len(indxLINES)):
            exprUPrimeIS[i] = LinExpr()
            for s in range(len(indxSHIFTS)):
                exprUPrimeIS[i].addTerms(1,uPrime_is[i][s])
            myModel.addConstr(exprUPrimeIS[i] == v_i[i])

        for i in range(len(indxLINES)):
            for s in range(len(indxSHIFTS)):
                if s == 0:
                    myModel.addConstr(uPrime_is[i][s] == u_is[i][s])
                else:
                    myModel.addConstr(uPrime_is[i][s] <= u_is[i][s])
                    myModel.addConstr(uPrime_is[i][s] >= u_is[i][s] - u_is[i][s-1])

        exprTSUPrime=len(indxLINES)*['']
        for i in range(len(indxLINES)):
            exprTSUPrime[i] = LinExpr()
            for s in range(len(indxSHIFTS)):
                exprTSUPrime[i].addTerms(tShiftStart[s],uPrime_is[i][s])
            myModel.addConstr(tStart_i[i] == exprTSUPrime[i])

        exprTSU=len(indxLINES)*['']
        for i in range(len(indxLINES)):
            exprTSU[i] = LinExpr()
            for s in range(len(indxSHIFTS)):
                exprTSU[i].addTerms(shiftsData[(indxSHIFTS[s],'maxTotalTime')],u_is[i][s])
            myModel.addConstr(tStop_i[i] == tStart_i[i] + exprTSU[i])

        for i in range(len(indxLINES)):
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    myModel.addConstr(t_ij[i][j] >= tStart_i[i] - PARAM_GLOBAL['bigMT'] * (1 - z_ij[i][j]))
                    myModel.addConstr(t_ij[i][j] >= 0)
                    myModel.addConstr(t_ij[i][j] <= PARAM_GLOBAL['bigMT'] * z_ij[i][j])
                    try:
                        prodtime_ij = prodTime[(ITEMS_ALL[j].material,indxLINES[i])]
                    except:
                        print('WARNING: brak czasu produkcji',',',j,',',i,',',ITEMS_ALL[j].material,',',indxLINES[i],',',prodtime_ij)
                        prodtime_ij = 0
                    finally:
                        myModel.addConstr(t_ij[i][j] <= tStop_i[i] - x_ij[i][j] * prodtime_ij)

        for i in range(len(indxLINES)):
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                     for k in range(len(indxJLong)):
                        if (ITEMS_ALL[k].material,indxLINES[i]) in EFFICIENCY:
                            try:
                                prodtime_ij = prodTime[(ITEMS_ALL[j].material,indxLINES[i])]
                                changeover_ijk = MP[(ITEMS_ALL[j].material,ITEMS_ALL[k].material,indxLINES[i])]
                            except:
                                print('WARNING: brak czasu produkcji',',',j,',',i,',',ITEMS_ALL[j].material,',',indxLINES[i],',',prodtime_ij,', LUB ')
                                print('brak czasu przezbrojenia',',',j,',',k,',',i,',',ITEMS_ALL[j].material,',',ITEMS_ALL[k].material,',',indxLINES[i],',',changeover_ijk)
                                prodtime_ij = 0
                                changeover_ijk = 0
                            finally:
                                myModel.addConstr(t_ij[i][k] - t_ij[i][j] >= - (1 - y_ijk[i][j][k]) * PARAM_GLOBAL['bigMT'] + x_ij[i][j] * prodtime_ij + y_ijk[i][j][k] * changeover_ijk)

        exprTij = len(indxJLong)*['']
        for j in range(len(indxJLong)):
            if ITEMS_ALL[j].material in matIndxMix:
                if ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm1':
                    if ITEMS_ALL[j+1].nbr[-2:len(ITEMS_ALL[j+1].nbr)] == 'm2': # podwójne zabezpieczenie
                        if ITEMS_ALL[j+1].nbr[0:len(ITEMS_ALL[j+1].nbr)-2] == ITEMS_ALL[j].nbr[0:len(ITEMS_ALL[j].nbr)-2]: # a nawet potrójne
                            exprTij[j] = LinExpr()
                            for i in range(len(indxLINES)):
                                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                                    if (ITEMS_ALL[j+1].material,indxLINES[i]) in EFFICIENCY:
                                        exprTij[j].addTerms(1,t_ij[i][j])
                                        exprTij[j].addTerms(-1,t_ij[i][j+1])
                            myModel.addConstr(exprTij[j] <= PARAM_GLOBAL['deltaMIX'] + PARAM_GLOBAL['bigMT'] * a_j[j])
                            myModel.addConstr(exprTij[j] >= - PARAM_GLOBAL['deltaMIX'] - PARAM_GLOBAL['bigMT'] * a_j[j])

        exprXijProp = len(indxJLong)*['']
        for j in range(len(indxJLong)):
            if ITEMS_ALL[j].material in matIndxMix:
                if ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm1':
                    if ITEMS_ALL[j+1].nbr[-2:len(ITEMS_ALL[j+1].nbr)] == 'm2':
                        if ITEMS_ALL[j+1].nbr[0:len(ITEMS_ALL[j+1].nbr)-2] == ITEMS_ALL[j].nbr[0:len(ITEMS_ALL[j].nbr)-2]:
                            exprXijProp[j] = LinExpr()
                            for i in range(len(indxLINES)):
                                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                                    if (ITEMS_ALL[j+1].material,indxLINES[i]) in EFFICIENCY:
                                        exprXijProp[j].addTerms(MAT_MIX[matIndxMix.index(ITEMS_ALL[j].material)].udzial2,x_ij[i][j])
                                        exprXijProp[j].addTerms(-MAT_MIX[matIndxMix.index(ITEMS_ALL[j].material)].udzial1,x_ij[i][j+1])
                            myModel.addConstr(exprXijProp[j] == 0)

        exprSumB = len(indxJLong)*['']
        for j in range(len(indxJLong)):
            if ITEMS_ALL[j].material in matIndxMix:
                if ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm1':
                    if ITEMS_ALL[j+1].nbr[-2:len(ITEMS_ALL[j+1].nbr)] == 'm2':
                        if ITEMS_ALL[j+1].nbr[0:len(ITEMS_ALL[j+1].nbr)-2] == ITEMS_ALL[j].nbr[0:len(ITEMS_ALL[j].nbr)-2]:
                            exprSumB[j] = LinExpr()
                            for i in range(len(indxLINES)):
                                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                                    if (ITEMS_ALL[j+1].material,indxLINES[i]) in EFFICIENCY:
                                        myModel.addConstr(z_ij[i][j] + z_ij[i][j+1] <= b_ij[i][j] + 1)
                                        myModel.addConstr(z_ij[i][j] + z_ij[i][j+1] >= 2 * b_ij[i][j])
                                        exprSumB[j].addTerms(1,b_ij[i][j])
                            myModel.addConstr(a_j[j] == exprSumB[j])

        for j in range(len(indxJLong)):
            if ITEMS_ALL[j].material in matIndxMix:
                if ITEMS_ALL[j].nbr[-2:len(ITEMS_ALL[j].nbr)] == 'm1':
                    if ITEMS_ALL[j+1].nbr[-2:len(ITEMS_ALL[j+1].nbr)] == 'm2':
                        if ITEMS_ALL[j+1].nbr[0:len(ITEMS_ALL[j+1].nbr)-2] == ITEMS_ALL[j].nbr[0:len(ITEMS_ALL[j].nbr)-2]:
                            myModel.addConstr(q_j[j] == q_j[j+1])

        exprZijMix = len(indxJLong)*['']
        for j in range(len(indxJLong)):
            if ITEMS_ALL[j].material in matIndxMix:
                exprZijMix[j] = LinExpr()
                for i in range(len(indxLINES)):
                    if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                        exprZijMix[j].addTerms(1,z_ij[i][j])
                myModel.addConstr(exprZijMix[j] <= 1)

        myModel.Params.MIPGap = PARAM_GLOBAL['epsilon_SOLVER']
        myModel.Params.TimeLimit = PARAM_GLOBAL['timeLimitSeconds_SOLVER']
        myModel.optimize()

        rows = []
        for i in range(len(indxLINES)):
            if w_i[i].x > PARAM_GLOBAL['epsilon']:
                print(w_i[i].VarName,' = ',w_i[i].x)
                sStart = 0
                tStart = 0
                while u_is[i][sStart].x <= PARAM_GLOBAL['epsilon']:
                    tStart += shiftsData[(indxSHIFTS[sStart],'maxTotalTime')]
                    sStart += 1
                sStop = len(indxSHIFTS) - 1
                while u_is[i][sStop].x <= PARAM_GLOBAL['epsilon']:
                    sStop -= 1
            if v_i[i].x > PARAM_GLOBAL['epsilon']:
                print(v_i[i].VarName,' = ',v_i[i].x)
                print(tStart_i[i].VarName,' = ',tStart_i[i].x)
                print(tStop_i[i].VarName,' = ',tStop_i[i].x)
                rowItem = WYNIK(indxLINES[i],'przezbrojenia','n/d','n/d','n/d',T1_i[i].x,0, 'n/d','n/d','n/d','n/d','n/d','n/d','n/d')
                rows.append(rowItem)
            if n_i[i].x > PARAM_GLOBAL['epsilon']:
                print(n_i[i].VarName,' = ',n_i[i].x)
                print(n_i0[i].VarName,' = ',n_i0[i].x)
                kol = int(round(n_i[i].x)) * ['']
                tKol = int(round(n_i[i].x)) * ['']
                tKol[0] = tStart
            if T1_i[i].x > PARAM_GLOBAL['epsilon']:
                print(T1_i[i].VarName,' = ',T1_i[i].x)
            if T2_i[i].x > PARAM_GLOBAL['epsilon']:
                print(T2_i[i].VarName,' = ',T2_i[i].x)
            for s in range(len(indxSHIFTS)):
                if u_is[i][s].x > PARAM_GLOBAL['epsilon']:
                    print (u_is[i][s].VarName,' = ',u_is[i][s].x)
                if uPrime_is[i][s].x > 0:
                    print (uPrime_is[i][s].VarName,' = ',uPrime_is[i][s].x)
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    if p_ij[i][j].x > PARAM_GLOBAL['epsilon']:
                        #print("XXXXXXXXXXXXXXXXXXXXXXXXX p_ij[i][j].x =",p_ij[i][j].x,", i=", i, ", j=", j)
                        kol[int(round(p_ij[i][j].x))-1] = j
            for uu in range(1,int(n_i[i].x)):
                if (ITEMS_ALL[kol[uu]].material,indxLINES[i]) in EFFICIENCY:
                    print(str(kol[uu]) + ': ' + ITEMS_ALL[kol[uu]].nbr)
                    if p_ij[i][kol[uu]].x > PARAM_GLOBAL['epsilon']:
                        tKol[uu] = tKol[uu-1] + x_ij[i][kol[uu-1]].x*prodTime[(ITEMS_ALL[kol[uu-1]].material,indxLINES[i])]
                        tKol[uu] += MP[(ITEMS_ALL[kol[uu-1]].material,ITEMS_ALL[kol[uu]].material,indxLINES[i])]
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    if x_ij[i][j].x > PARAM_GLOBAL['epsilon']:
                        print(x_ij[i][j].VarName,' = ',x_ij[i][j].x)
                        line = indxLINES[i]
                        nbr = ITEMS_ALL[j].nbr
                        material = ITEMS_ALL[j].material
                        weight = x_ij[i][j].x
                        originalWeight = ITEMS_ALL[j].originalWeight
                        time = x_ij[i][j].x*prodTime[(ITEMS_ALL[j].material,indxLINES[i])]
                        pos = int(round(p_ij[i][j].x))
                        #rTStart = tKol[kol.index(j)] #to jest stary sposób, zostawiam dla ew. celów porównawczych
                        rTStart = t_ij[i][j].x
                        rTStop = rTStart + time
                        #alert wstawka tu wpisuję na stałe godzinę pauzy: 13:45-14:15 przy założeniu startu o 06:00, tj. po 465 minutach
                        if rTStart > 465:
                            rTStart += 30
                        if rTStop > 465:
                            rTStop += 30
                        #alert koniec wstawki
                        mtName = ITEMS_ALL[j].materialName
                        prLine = ITEMS_ALL[j].prefLine
                        grp = ITEMS_ALL[j].group
                        subset = ITEMS_ALL[j].indSubset
                        if subset =='J0':
                            subset = 'zaległe rozpoczęte (J0)'
                        elif subset =='J1':
                            subset = 'zaległe nierozpoczęte (J1)'
                        elif subset =='J2_1':
                            subset = 'produkcyjne niepodzielne (J2_1)'
                        elif subset =='J2_2':
                            subset = 'produkcyjne podzielne (J2_2)'
                        elif subset =='J3':
                            subset = 'planowane (J3)' 
                        dateR = ITEMS_ALL[j].date_r
                        rowItem = WYNIK(line,nbr,material,weight,originalWeight,time,pos,rTStart,rTStop,mtName,prLine,grp,dateR,subset)
                        rows.append(rowItem)
                    if z_ij[i][j].x > PARAM_GLOBAL['epsilon']:
                        print(z_ij[i][j].VarName,' = ',z_ij[i][j].x)
                    if p_ij[i][j].x > PARAM_GLOBAL['epsilon']:
                        print(p_ij[i][j].VarName,' = ',p_ij[i][j].x)
            
            for j in range(len(indxJLong)):
                if (ITEMS_ALL[j].material,indxLINES[i]) in EFFICIENCY:
                    if y_ijk[i][j][-1].x > PARAM_GLOBAL['epsilon']:
                        print(y_ijk[i][j][-1].VarName,' = ',y_ijk[i][j][-1].x)
                    if y_ijk[i][-1][j].x > PARAM_GLOBAL['epsilon']:
                        print(y_ijk[i][-1][j].VarName,' = ',y_ijk[i][-1][j].x)
                    for k in range(len(indxJLong)):
                        if (ITEMS_ALL[k].material,indxLINES[i]) in EFFICIENCY:
                            if y_ijk[i][j][k].x > PARAM_GLOBAL['epsilon']:
                                print(y_ijk[i][j][k].VarName,' = ',y_ijk[i][j][k].x)
        for j in range(len(indxJ)):
            if q_j[j].x > PARAM_GLOBAL['epsilon']:
                print(q_j[j].VarName, ' = ', q_j[j].x)
            else:
                if ITEMS_ALL[j].indSubset in ['J0','J1','J2_1','J2_2']:
                    line = 'xxx_nieprzypisane'
                    nbr = ITEMS_ALL[j].nbr
                    material = ITEMS_ALL[j].material
                    weight = 0
                    originalWeight = ITEMS_ALL[j].originalWeight
                    time = 0
                    pos = 'n/d'
                    rTStart = 'n/d'
                    rTStop = 'n/d'
                    mtName = ITEMS_ALL[j].materialName
                    prLine = ITEMS_ALL[j].prefLine
                    grp = ITEMS_ALL[j].group
                    subset = ITEMS_ALL[j].indSubset
                    if subset =='J0':
                        subset = 'zaległe rozpoczęte (J0)'
                    elif subset =='J1':
                        subset = 'zaległe nierozpoczęte (J1)'
                    elif subset =='J2_1':
                        subset = 'produkcyjne niepodzielne (J2_1)'
                    elif subset =='J2_2':
                        subset = 'produkcyjne podzielne (J2_2)'
                    elif subset =='J3':
                        subset = 'planowane (J3)' 
                    dateR = ITEMS_ALL[j].date_r
                    rowItem = WYNIK(line,nbr,material,weight,originalWeight,time,pos,rTStart,rTStop,mtName,prLine,grp,dateR,subset)
                    rows.append(rowItem)

        sRows = sorted(rows, key = lambda x: (x.line, x.order))

        tt=PARAM_GLOBAL['godzina_start']

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'WYNIKI'
        sheet.cell(row = 1, column = 1).value = 'LINIA'
        sheet.cell(row = 1, column = 2).value = 'ZLECENIE'
        sheet.cell(row = 1, column = 3).value = 'MATERIAŁ'
        sheet.cell(row = 1, column = 4).value = 'NAZWA'
        sheet.cell(row = 1, column = 5).value = 'LINIA PRIORYTETOWA'
        sheet.cell(row = 1, column = 6).value = 'GRUPA'
        sheet.cell(row = 1, column = 7).value = 'DATA'
        sheet.cell(row = 1, column = 8).value = 'KATEGORIA'
        sheet.cell(row = 1, column = 9).value = 'WAGA'
        sheet.cell(row = 1, column = 10).value = 'ORYGINALNA WAGA'
        sheet.cell(row = 1, column = 11).value = 'CZAS M'
        sheet.cell(row = 1, column = 12).value = 'CZAS H'
        sheet.cell(row = 1, column = 13).value = 'CZAS H:M'
        sheet.cell(row = 1, column = 14).value = 'KOLEJNOŚĆ'
        sheet.cell(row = 1, column = 15).value = 'START'
        sheet.cell(row = 1, column = 16).value = 'STOP'
        
        fields_to_csv = ['LINIA','ZLECENIE','MATERIAL','NAZWA','LINIA PRIORYTETOWA','GRUPA','DATA','KATEGORIA','WAGA','ORYGINALNA WAGA',
          'CZAS M','CZAS H','CZAS H:M','KOLEJNOSC','START','STOP']
        rows_to_csv =[]

        for i in range(len(rows)):
            sheet.cell(row = i+2, column = 1).value = sRows[i].line
            sheet.cell(row = i+2, column = 2).value = sRows[i].nbr
            sheet.cell(row = i+2, column = 3).value = sRows[i].material
            sheet.cell(row = i+2, column = 4).value = sRows[i].materialName
            sheet.cell(row = i+2, column = 5).value = sRows[i].prefLine
            sheet.cell(row = i+2, column = 6).value = sRows[i].group
            sheet.cell(row = i+2, column = 7).value = sRows[i].date_r
            sheet.cell(row = i+2, column = 8).value = sRows[i].indSubset
            sheet.cell(row = i+2, column = 9).value = sRows[i].weight
            sheet.cell(row = i+2, column = 10).value = sRows[i].originalWeight
            sheet.cell(row = i+2, column = 11).value = int(round(sRows[i].time))
            sheet.cell(row = i+2, column = 12).value = round(sRows[i].time/60,2)
            sheet.cell(row = i+2, column = 13).value = '{:02d}:{:02d}'.format(*divmod(int(round(sRows[i].time)), 60))
            sheet.cell(row = i+2, column = 14).value = sRows[i].order
            if sRows[i].realTimeStart != 'n/d':
                tttt = 60*PARAM_GLOBAL['godzina_start'].hour+PARAM_GLOBAL['godzina_start'].minute
                sheet.cell(row = i+2, column = 15).value = '{:02d}:{:02d}'.format(*divmod(int(round(sRows[i].realTimeStart+tttt)), 60))
                sheet.cell(row = i+2, column = 16).value = '{:02d}:{:02d}'.format(*divmod(int(round(sRows[i].realTimeStop+tttt)), 60))
                rows_to_csv.append([sRows[i].line, sRows[i].nbr, sRows[i].material, sRows[i].materialName, sRows[i].prefLine,
                            sRows[i].group, sRows[i].date_r, sRows[i].indSubset, sRows[i].weight,
                            sRows[i].originalWeight,
                            int(round(sRows[i].time)), round(sRows[i].time / 60, 2),
                            '{:02d}:{:02d}'.format(*divmod(int(round(sRows[i].time)), 60)), sRows[i].order,
                            '{:02d}:{:02d}'.format(*divmod(int(round(sRows[i].realTimeStart + tttt)), 60)),
                            '{:02d}:{:02d}'.format(*divmod(int(round(sRows[i].realTimeStop + tttt)), 60)) ] )

            else:
                rows_to_csv.append([sRows[i].line,sRows[i].nbr,sRows[i].material,sRows[i].materialName,sRows[i].prefLine,
                        sRows[i].group,sRows[i].date_r,sRows[i].indSubset,sRows[i].weight,sRows[i].originalWeight,
                        int(round(sRows[i].time)), round(sRows[i].time/60,2),
                        '{:02d}:{:02d}'.format(*divmod(int(round(sRows[i].time)), 60)), sRows[i].order,'',''])
        wb.save(PATHS_TO_DATA['wyniki'])
        with open('anx_wynik.csv', 'w', encoding='utf-8') as f:
            # using csv.writer method from CSV package
            write = csv.writer(f)
            write.writerow(fields_to_csv)
            write.writerows(rows_to_csv)
                                       
        print('wersja4 nowa')
env.dispose()   