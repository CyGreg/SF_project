import csv
import sys
import datetime
import openpyxl
from datetime import date
import gurobipy as gp
from gurobipy import GRB
import holidays
import calendar


#pyinstaller --onefile --add-binary "venv/Lib/site-packages/cvxopt/.lib/libopenblas.dll;." main.py
#pyinstaller -F -w --add-binary "C:\Users\az\PycharmProjects\pythonProject\venv/Lib/site-packages/cvxopt/.lib/libopenblas.dll;." main.py

END_DATE = datetime.datetime.strptime('31.07.2044', "%d.%m.%Y")
TODAY = datetime.datetime.now()
#TODAY = datetime.datetime.strptime('12.09.2023', "%d.%m.%Y")

def nbr_of_weeks(year):
    date = datetime.datetime.strptime('31.12.'+str(year), "%d.%m.%Y")
    return int(date.strftime('%W'))
def nbr_of_days(year):
    date = datetime.datetime.strptime('31.12.'+str(year), "%d.%m.%Y")
    return int(date.strftime('%j'))

def previous_working_day(check_day_,more_holidays):
    offset = max(1, (check_day_.weekday() + 6) % 7 - 3)
    most_recent = check_day_ - datetime.timedelta(offset)
    if (most_recent not in holidays.PL()) and (most_recent not in more_holidays):
        return most_recent
    else:
        return previous_working_day(most_recent, more_holidays)

def next_working_day(check_day_,more_holidays):

    nwd = check_day_ + datetime.timedelta(1)
    if (nwd.weekday() < 5) and (nwd not in holidays.PL()) and (nwd not in more_holidays):
        return nwd
    else:
        return next_working_day(nwd, more_holidays)

def load_options(wb):
    sh = wb['opcje']
    tm_limit_w = sh.cell(row=4,column=3).value
    tm_limit_d = sh.cell(row=4, column=4).value
    backlogs_w = sh.cell(row=3, column=3).value
    backlogs_d = sh.cell(row=3, column=4).value

    return tm_limit_w, tm_limit_d, backlogs_w, backlogs_d



def load_parameters(wb):

    sh = wb['input']
    path_w = sh.cell(row=5,column=3).value
    path_d = sh.cell(row=7,column=3).value
    path_eff = sh.cell(row=9,column=3).value
    tps_path = sh.cell(row=11, column=3).value
    nbr_of_shifts = sh.cell(row=5, column=6).value
    nbr_working_days = sh.cell(row=7, column=6).value
    capacity = sh.cell(row=9, column=6).value
    nbr_hours= sh.cell(row=11, column=6).value
    horizon_w = [sh.cell(row=14,column=6).value,sh.cell(row=14,column=7).value]
    horizon_d = [sh.cell(row=16,column=6).value,sh.cell(row=16,column=7).value]
    txt = str(sh.cell(row=14, column=3).value)
    txt = txt.replace(' ','')
    plastrownia = txt.split(';')


    return nbr_of_shifts, capacity, nbr_working_days, path_w, path_d, path_eff, nbr_hours, tps_path, horizon_w, horizon_d, plastrownia


def load_exceptions_calendar(wb, horizon_w, horizon_d):

    print()
    print('------------------------------ EXCEPTIONS -------------------------------')
    print('CALENDAR EXCEPTIONS: ')
    sh = wb['wyjątki kalendarz']
    licznik = 2
    rok_b = TODAY.year
    max_row = sh.max_row

    for l in LINES:
        for t in range(ZAKRES[0][0],ZAKRES[0][1]+1):
            CPTY[0][(l,t)] = NBR_OF_SHIFTS * NBR_WORKING_DAYS

    while True:
        licznik += 1
        rok = sh.cell(row = licznik, column = 2).value

        if licznik >= min(1000,max_row+5):
            break
        if rok in [None,0,'']:
            continue


        try:
            nr_tyg = (rok - rok_b) * nbr_of_weeks(rok_b) + sh.cell(row = licznik, column = 3).value
            liczba_zm = int(sh.cell(row = licznik, column = 4).value)
            dni_rob = int(sh.cell(row = licznik, column = 5).value)
            zalogi = int(sh.cell(row=licznik, column=6).value)

        except ValueError:
            continue

        if nr_tyg >= ZAKRES[0][0] and nr_tyg <= ZAKRES[0][1]:
            MAX_L[0][nr_tyg] = dni_rob * zalogi * liczba_zm
            for l in LINES:
                CPTY[0][(l, nr_tyg)] = dni_rob * liczba_zm
            print('tydzien = ', nr_tyg, ', moce prod. (plastry) = ', MAX_L[0][nr_tyg], 'zmian')

    for l in LINES:
        for t in range(ZAKRES[1][0],ZAKRES[1][1]+1):
            CPTY[1][(l,t)] = NBR_OF_SHIFTS

    licznik = 2
    while True:
        licznik += 1
        data = sh.cell(row = licznik, column = 8).value
        if licznik >= min(1000,max_row+5):
            break
        if data in [None,0,'']:
            continue


        try:
            nr_dnia = (data.year - rok_b) * nbr_of_days(rok_b) + int(data.strftime('%j'))
            liczba_zm = int(sh.cell(row = licznik, column = 9).value)
            zalogi = int(sh.cell(row=licznik, column=10).value)
            powtarzaj = str(sh.cell(row=licznik, column=11).value)

        except ValueError:
            continue

        if powtarzaj == 'tak':
            for t in range(ZAKRES[1][0],ZAKRES[1][1]+1):
                if (t - nr_dnia) % 7 == 0:
                    MAX_L[1][t] = zalogi * liczba_zm
                    for l in LINES:
                        CPTY[1][(l,t)] = liczba_zm
                    print('dzien = ', str(data+datetime.timedelta(t-nr_dnia)), '(',t, '), moce prod. (plastry) = ', MAX_L[1][t], 'zmian')
        else:
            if nr_dnia >= ZAKRES[1][0] and nr_dnia <= ZAKRES[1][1]:
                MAX_L[1][nr_dnia] = zalogi * liczba_zm
                for l in LINES:
                    CPTY[1][(l,nr_dnia)] = liczba_zm
                print('dzien = ', str(data), '(', nr_dnia, '), moce prod. (plastry) = ', MAX_L[1][nr_dnia], 'zmian')



    print()


def load_exceptions_lines(wb, horizon_w, horizon_d):

    print('LINES EXCEPTIONS: ')

    sh = wb['wyjątki linie']
    licznik = 2
    rok_b = TODAY.year
    max_row = sh.max_row

    while True:
        licznik += 1
        rok = sh.cell(row = licznik, column = 2).value
        if licznik >= min(1000,max_row+5):
            break
        if rok in [None,0,'']:
            continue

        try:
            nr_tyg = (rok - rok_b) * nbr_of_weeks(rok_b) + sh.cell(row = licznik, column = 3).value
            dost = int(sh.cell(row=licznik, column=5).value)
            linie_txt = str(sh.cell(row=licznik, column=4).value).upper()
            linie_txt = linie_txt.replace(' ','')
            linie_exc = linie_txt.split(';')

        except KeyError:
            continue

        if nr_tyg < ZAKRES[0][0] or nr_tyg > ZAKRES[0][1]:
            continue

        for l in linie_exc:
            MAX_L[0][nr_tyg] += max(0, CPTY[0][(l, nr_tyg)] * (dost - 100)/100)     #zmiana 28.11
            CPTY[0][(l,nr_tyg)] = CPTY[0][(l,nr_tyg)] * dost / 100                  #zmiana 28.11
            print('linia = ',l,', tydzien = ', nr_tyg, ', dostepnosc = ', CPTY[0][(l,nr_tyg)], 'zmian')

    licznik = 2
    while True:
        licznik += 1
        data = sh.cell(row=licznik, column=8).value
        if licznik >= min(1000,max_row+5):
            break
        if data in [None, 0, '']:
            continue

        try:
            nr_dnia = (data.year - rok_b) * nbr_of_days(rok_b) + int(data.strftime('%j'))
            dost = int(sh.cell(row=licznik, column=10).value)
            linie_txt = str(sh.cell(row=licznik, column=9).value).upper()
            linie_txt = linie_txt.replace(' ', '')
            linie_exc = linie_txt.split(';')

        except ValueError:
            continue

        if nr_dnia < ZAKRES[1][0] or nr_dnia > ZAKRES[1][1]:
            continue

        for l in linie_exc:
            MAX_L[1][nr_dnia] += max(0, CPTY[1][(l, nr_dnia)] * (dost - 100) / 100)     #zmiana 28.11
            CPTY[1][(l, nr_dnia)] = CPTY[1][(l, nr_dnia)] * dost / 100                  #zmiana 28.11
            print('linia = ', l, ', dzien = ', str(data), '(', nr_dnia, '), dostepnosc = ', CPTY[1][(l, nr_dnia)], 'zmian')

    rows = []                                                             #zmiana 28.11 (przeniesione z funkcji load_exception_calendar)
    for t in range(ZAKRES[0][0], ZAKRES[0][1] + 1):
        rows.append([t, WEEK_DATE[t], MAX_L[0][t] * NBR_HOURS, 'W'])
    for t in range(ZAKRES[1][0], ZAKRES[1][1] + 1):
        rows.append([t, str(DAY_DATE[t])[0:10], MAX_L[1][t] * NBR_HOURS, 'D'])
        #print(t, MAX_L[1][t])

    with open('AMX_capacity.csv', 'w',encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['nr okresu', 'okres', 'limit', 'W/D'])
        write.writerows(rows)

    print('----------------------------------------------------------------------------')
    return 1


def load_data_from_excel(wb_name,sh_name):      # zwraca wiersze excela jako listę słowników (klucze każdego słownika to pierwszy wiersz excela)
    wb = openpyxl.load_workbook(wb_name, data_only=True)
    sh = wb[sh_name]
    max_row = sh.max_row
    max_col = sh.max_column
    fields = []

    for j in range(1,max_col+1):
        fields.append(sh.cell(row = 1,column = j).value)

    rows = []
    for i in range(2, max_row + 1):
        row = {}

        for j in range(1, max_col + 1):
            row[fields[j - 1]] = sh.cell(row=i, column=j).value
        rows.append(row)

    return rows


def edit_date_format(zakres):

    day_date = {}
    week_date = {}

    for t in range(zakres[0][1]-zakres[0][0]+1):
        data = HORIZON_W[0] + datetime.timedelta(7*t)
        t_0 = ZAKRES[0][0]
        w = data.strftime('%W')
        week_date[t_0+t] = str(data.year) + 'T' + w

    for t in range(zakres[1][1] - zakres[1][0] + 1):
        data = HORIZON_D[0] + datetime.timedelta(t)
        t_0 = ZAKRES[1][0]
        day_date[t_0+t] = data

    return day_date, week_date


def load_efficiency(wb_name):
    wb = openpyxl.load_workbook(wb_name, data_only=True)
    sh = wb['wydajności']
    eff = {}
    lines = []
    products = []
    group = {}
    line_priority = {}
    short_text = {}

    licznik = 2
    while True:
        licznik += 1

        prod = str(sh.cell(row=licznik, column = 17).value)
        if prod in [None,'None','']:
            break
        if prod not in products:
            products.append(prod)

        group[prod] = str(sh.cell(row=licznik, column=22).value)
        short_text[prod] = str(sh.cell(row = licznik, column = 18).value)
        if group[prod] in ['','None',None]:
            BRAKI.append([prod,'brak grupy produktowej'])

        line_priority[prod] = str(sh.cell(row=licznik, column=39).value)
        if line_priority[prod] in ['','None',None]:
            BRAKI.append([prod, 'brak informacji o linii priorytetowej'])


        for j in range(23,38):

            line = str(sh.cell(row=2,column=j).value)
            if line not in lines:
                lines.append(line)

            try:
                eff[(prod,line)] = 8 * sh.cell(row=licznik,column=j).value
            except TypeError:
                eff[(prod,line)] = 0

    """
    rows = []
    for prod in products:
        for l in lines:
            rows.append([prod,l,eff.get((prod,l),'brak')])

    with open('efficiency.csv', 'w') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['produkt','linia','wydajnosc'])
        write.writerows(rows)
    """

    return eff, lines, group, line_priority, short_text

def edit_sales_orders(sales,iter):
    sales_orders = {}
    products = []
    week_date = {}
    day_date = {}

    min_day_nbr = 1000
    min_week_nbr = 1000
    max_day_nbr = 0
    max_week_nbr = 0

    for row in sales:
        prod = str( row['Nr materiału'] )
        if prod in ['None','']:
            continue


        date2 = row['Bazowy termin rozp.']

        try:
            year2 = int(date2.strftime('%Y'))
        except AttributeError:
            date2 = datetime.datetime.strptime(date2, "%d.%m.%Y")


        year2 = int(date2.strftime('%Y'))
        week2 = int(date2.strftime('%W'))
        day2 = int(date2.strftime('%j'))


        year1 = int(TODAY.strftime('%Y'))
        week1 = int(TODAY.strftime('%W'))
        day1 = int(TODAY.strftime('%j'))

        day_nbr = int(date2.strftime('%j')) + (year2 - year1) * nbr_of_days(year1)
        week_nbr = int(date2.strftime('%W')) + (year2 - year1) * nbr_of_weeks(year1)
        max_day_nbr = max(max_day_nbr, day_nbr)
        min_day_nbr = min(min_day_nbr, max(day_nbr, day1))
        max_week_nbr = max(max_week_nbr, week_nbr)
        min_week_nbr = min(min_week_nbr, max(week_nbr, week1))
        day_date[day_nbr] = date2 #datetime.strptime(date2, '%d.%m.%Y')
        week_date[week_nbr] = str(year2) + 'T' + str(week2)

        if iter == 0:
            if week_nbr >= ZAKRES[iter][0] and week_nbr <= ZAKRES[iter][1]:
                sales_orders[(prod, week_nbr)] = sales_orders.get((prod,week_nbr),0) + float(row['waga'])
        else:
            if day_nbr >= ZAKRES[iter][0] and day_nbr <= ZAKRES[iter][1]:
                sales_orders[(prod, day_nbr)] = sales_orders.get((prod,day_nbr),0) + float(row['waga'])

        if prod not in products:
            products.append(prod)

    rows = []

    for t in range(ZAKRES[iter][0],ZAKRES[iter][1]+1):
        if iter == 0:
            rows.append([0,t,WEEK_DATE[t],0])
        else:
            rows.append([0,t,str(DAY_DATE[t])[0:10],0])
    for mat in sales_orders:
        if iter == 0:
            rows.append([mat[0],mat[1],WEEK_DATE[mat[1]], sales_orders[mat]])
        else:
            rows.append([mat[0], mat[1], str(DAY_DATE[mat[1]])[0:10], sales_orders[mat]])

    with open('sales_orders'+str(iter)+'.csv', 'w',encoding = 'utf-8') as f:
        write = csv.writer(f)
        write.writerow(['Material','nr okresu','okres','waga'])
        write.writerows(rows)

    if iter == 0:
        return sales_orders, products
    else:
        return sales_orders, products


def edit_tps(czasy_tps,products,limit):
    tps = {}
    for row in czasy_tps:
        tps[str(row['Materiał'])] = min(int(row['dni na wyrotowanie']),limit)

    for prod in products:
        try:
            t = tps[prod]
        except KeyError:
            BRAKI.append([prod, 'brak TPS'])

    return tps


def check_production_version(products,lines,efficiency):
    braki = []
    for prod in products:
        eff_pr = 0
        for l in lines:

            a = efficiency.get((prod, l), None)
            if a == None:
                a = 0
            eff_pr += a
        if eff_pr == 0:
            braki.append([prod,'brak wersji produkcyjnej'])
            products.remove(prod)

    return braki

def save_braki():
    with open('erp_AMX_braki.csv', 'w',encoding = 'utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['material', 'braki danych'])
        write.writerows(BRAKI)


def create_coeff_x(line_priority):
    coeff_x = {}

    for prod in line_priority:
        l = line_priority[prod]
        coeff_x[(prod,l)] = 1

    return coeff_x

def create_variables(m,zakres,products,lines,efficiency):

    x = {}
    s = {}
    z = {}
    y = {}
    yb1 = {}
    yb2 = {}
    yb3 = {}

    for prod in products:

        for t in range(zakres[0],zakres[1]+1):
            x[(prod,t)] = m.addVar(lb = 0, vtype=GRB.CONTINUOUS)
            s[(prod,t)] = m.addVar(lb = 0, vtype=GRB.CONTINUOUS)
            z[(prod,t)] = m.addVar(lb = 0, vtype=GRB.CONTINUOUS)
            for l in lines:
                if efficiency.get((prod,l),None) not in [None,0]:
                    x[(prod,l,t)] = m.addVar(lb = 0, vtype=GRB.CONTINUOUS)

        s[(prod, zakres[0] - 1)] = m.addVar(lb = 0, vtype=GRB.CONTINUOUS)
        z[(prod, zakres[0] - 1)] = m.addVar(lb = 0, vtype=GRB.CONTINUOUS)

    for l in lines:
        for t in range(zakres[0], zakres[1] + 1):
            y[(l,t)] = m.addVar(lb = 0,  vtype=GRB.CONTINUOUS)                  #ilość zmian pracy linii l w tygodniu t
            yb1[(l,t)] = m.addVar(lb = 0, vtype=GRB.CONTINUOUS)                 # dodatkowe moce [il zmian/tydz.]

    for t in range(zakres[0], zakres[1] + 1):
        yb2[t] = m.addVar(lb=0, vtype=GRB.CONTINUOUS)                           # będzie odpowiadać za max przekroczenia
                                                                                # mocy jednej linii w tygodniu
        yb3[t] = m.addVar(lb=0, vtype=GRB.CONTINUOUS)                           # będzie odpowiadać za przekroczenie sumy mocy
                                                                                # w jednym tygodniu


    return x,s,z,y,yb1,yb2,yb3




def create_balance_cns(m,yb1,yb2,yb3, x,y,z,s,zakres,products,lines,packing_lines,sales_orders,eff,util_rate,max_l,nbr_of_shifts,nbr_working_days,cpty):

    t0 = zakres[0] - 1
    t_end = zakres[1]

    for prod in products:
        m.addConstr(s[(prod, t0)] == 0)
        m.addConstr(z[(prod, t0)] == 0)
        m.addConstr(z[(prod, t_end)] == 0)
        m.addConstr(s[(prod, t_end)] == 0)

        for t in range(zakres[0], zakres[1] + 1):
            d = sales_orders.get( (prod,t), 0 )
            if BACKLOGS_W == 'tak':
                m.addConstr( x[(prod, t)] + s[(prod, t - 1)] - z[(prod,t-1)] == d + s[(prod, t)] - z[(prod,t)])
            else:
                m.addConstr( x[(prod, t)] + s[(prod, t - 1)]  == d + s[(prod, t)] )

            m.addConstr( sum(x[(prod,l,t)] for l in lines if eff.get((prod,l),0) > 0) == x[(prod,t)] )
            m.addConstr( s[(prod,t-1)] <= d )

    for l in lines:
        for t in range(zakres[0], zakres[1] + 1):
            m.addConstr( sum(x[(prod,l,t)]/eff[(prod,l)] for prod in products if eff.get((prod,l),0) > 0) <= util_rate * y[(l,t)] )

            if cpty[(l,t)] < 0.01:
                m.addConstr(y[(l, t)] == 0)
                m.addConstr(y[(l, t)] == 0)
            else:
                m.addConstr(y[(l, t)] <= cpty[(l,t)] + yb1[(l, t)])
                m.addConstr(y[(l, t)] <= cpty[(l,t)] + yb2[t])

    for t in range(zakres[0], zakres[1] + 1):
        m.addConstr( sum(y[(l,t)] for l in packing_lines) <= util_rate*max_l[t] + yb3[t] )


    return 1

def create_balance_cns2(m,yb1,yb2,yb3, x,y,z,s,zakres,products,lines,packing_lines,sales_orders,eff,util_rate,max_l, \
                        nbr_of_shifts,cpty,tps, group):

    err = {}
    era = {}
    f = {}
    g = {}
    xx = {}
    gx = {}    # wskazuje czy na linii l w okresie t jest coś z grupy gk, gk in ['WP','WL','WA','DR']

    t0 = zakres[0] - 1
    t_end = zakres[1]

    for prod in products:
        d_total = sum( sales_orders.get( (prod,t), 0 ) for t in range(zakres[0], zakres[1] + 1))
        for l in lines:
            if eff[(prod,l)] not in [None,0]:
                for t in range(zakres[0], zakres[1] + 1):
                    xx[(prod,l,t)] = m.addVar(vtype = GRB.BINARY)
                    m.addConstr(d_total*xx[(prod,l,t)] >= x[(prod,l,t)])

    for prod in products:
        m.addConstr( s[(prod, t0)] == 0)
        m.addConstr( z[(prod, t0)] == 0 )
        m.addConstr( z[(prod, t_end)] == 0 )
        m.addConstr( s[(prod, t_end)] == 0)

        for t in range(zakres[0], zakres[1] + 1):
            d = sales_orders.get( (prod,t), 0 )
            if BACKLOGS_D == 'tak':
                m.addConstr( x[(prod, t)] + s[(prod, t - 1)] - z[(prod,t-1)] == d + s[(prod, t)] - z[(prod,t)])
            else:
                m.addConstr(x[(prod, t)] + s[(prod, t - 1)]  == d + s[(prod, t)] )

            m.addConstr( sum(x[(prod,l,t)] for l in lines if eff.get((prod,l),0) > 0) == x[(prod,t)] )
            m.addConstr( s[(prod,t)] <= sum(sales_orders.get( (prod, q), 0 ) for q in range(t,t+tps.get(prod,7)-1)))

    for l in lines:
        for t in range(zakres[0], zakres[1] + 1):
            err[(l, t)] = m.addVar(lb=-100, vtype=GRB.CONTINUOUS)  #zmiana 28.11
            if cpty[(l, t)] == 2:   #zmiana 28.11
                m.addConstr( sum(x[(prod, l, t)] / eff[(prod, l)] for prod in products if eff.get((prod, l), 0) > 0) == util_rate * y[(l,t)] + err[(l,t)])
            m.addConstr( sum(x[(prod, l, t)] / eff[(prod, l)] for prod in products if eff.get((prod, l), 0) > 0) <= cpty[(l,t)] * util_rate + yb1[(l, t)])
            m.addConstr( sum(x[(prod, l, t)] / eff[(prod, l)] for prod in products if eff.get((prod, l), 0) > 0) <= (cpty[(l,t)]+1) * util_rate + yb2[t])
            m.addConstr( sum(x[(prod, l, t)] / eff[(prod, l)] for prod in products if eff.get((prod, l), 0) > 0) <= 3 * util_rate + yb2[t]/100)
            if cpty[(l,t)] == 0:
                m.addConstr(yb1[(l, t)] == 0) #zmiana 28.11

            era[(l, t)] = m.addVar(lb=0, vtype=GRB.INTEGER)
            f[(l, t)] = m.addVar(lb=0, vtype=GRB.INTEGER)
            g[(l, t)] = m.addVar(vtype = GRB.BINARY)
            m.addConstr( y[(l, t)] ==  f[(l, t)] )
            m.addConstr( f[(l, t)] <= 2 )
            m.addConstr( f[(l, t)] <= cpty[(l, t)] * g[(l, t)] )


            if t != zakres[1] + 100:
                m.addConstr( era[(l,t)] >= err[(l,t)] ) #(zakres[1]-zakres[0])/(t+1-zakres[0]) * err[(l,t)]  )
                m.addConstr( era[(l,t)] >= -err[(l,t)] ) #(zakres[1]-zakres[0])/(t+1-zakres[0])*(-err[(l,t)]) )


    for t in range(zakres[0], zakres[1] + 1):
        #m.addConstr( sum(f[(l,t)] + err[(l,t)] for l in packing_lines) <= max_l[t] + yb3[t] )
        m.addConstr(sum(x[(prod, l, t)] / eff[(prod, l)] for prod in products for l in packing_lines if eff.get((prod, l),0) > 0)
                    <= util_rate*max_l[t] + yb3[t])
        #print(t,max_l[t])
        if max_l[t] == 0:
            m.addConstr(yb3[t] == 0)
        """
        if t == t0 + 1:
            m.addConstr(sum(x[(prod, l, t)] / eff[(prod, l)] for prod in products for l in packing_lines if eff[(prod, l)] > 0)
                >= util_rate * max_l[t])
        """
    group_kinds = ['WP','WŁ','WA','DR']

    for gk in group_kinds:
        q = sum(1 for prod in products if group[prod] == gk)
        for l in lines:
            for t in range(zakres[0],zakres[1] + 1):
                gx[(gk, l, t)] = m.addVar(vtype=GRB.BINARY)
                m.addConstr( sum(xx[(prod, l, t)] for prod in products \
                                 if (eff[(prod, l)] not in [None,0] and group[prod] == gk)) <= q * gx[(gk, l, t)])

    return era, err, f, g, xx, gx


def capacity_summary(zakres_d, zakres_w, x, eff):
    rows = []
    for t in range(zakres_d[0],zakres_d[1]+1):
        for l in LINES:
            try:
                prod_qtty = sum(x[1][(prod, l, t)].X for prod in PRODUCTS[1] if eff.get((prod,l),0) not in [None,0])
                prod_time = sum(x[1][(prod, l, t)].X/eff.get((prod, l),0) for prod in PRODUCTS[1] if eff.get((prod,l),0) not in [None,0])
            except KeyError:
                prod_qtty = 0
                prod_time = 0

            o = str(DAY_DATE[t])[0:10]

            rows.append([t, l, CPTY[1][(l, t)], CPTY[1][(l, t)] - prod_time/UTIL_RATE, 'zmiana', 'D', o])
            rows.append([t, l, UTIL_RATE*8*CPTY[1][(l, t)], UTIL_RATE*8*CPTY[1][(l, t)] - 8*prod_time, 'h', 'D', o])

    for t in range(zakres_w[0],zakres_w[1]+1):
        for l in LINES:
            try:
                prod_qtty = sum(x[0][(prod, l, t)].X for prod in PRODUCTS[0] if eff.get((prod,l),0) not in [None,0])
                prod_time = sum(x[0][(prod, l, t)].X/eff.get((prod, l),0) for prod in PRODUCTS[0] if eff.get((prod,l),0) not in [None,0])
            except KeyError:
                prod_qtty = 0
                prod_time = 0

            o = str(WEEK_DATE[t])[0:7]

            rows.append([t, l, CPTY[0][(l, t)], CPTY[0][(l, t)] - prod_time/UTIL_RATE, 'zmiana', 'W', o])
            rows.append([t, l, UTIL_RATE*8*CPTY[0][(l, t)], UTIL_RATE*8*CPTY[0][(l, t)] - 8*prod_time, 'h', 'W', o])

    with open('capacity_summary.csv', 'w', encoding='utf-8') as f:
        # using csv.writer method from CSV package
        write = csv.writer(f)
        write.writerow(['nr okresu', 'linia', 'dostepnosc','wolne moce','jdn','D/W','okres'])
        write.writerows(rows)

    return 1


#MATERIAL_SPECIFICATION = load_data_from_excel('','Sheet1')
#PRODUCTION_VERSIONS = load_data_from_excel('','Sheet1')
#STOCK_LEVELS = load_data_from_excel('','Sheet1')
#COMPONENTS, CONSUMPTION, PARENT = load_structures(MATERIAL_SPECIFICATION)
ZAKRES = [0,0]
PATH = [0,0]
M = [0,0]
SALES_ORDERS = [0,0]
PRODUCTS = [0,0]
SALES = [0,0]
X = [0,0]
S = [0,0]
Z = [0,0]
Y = [0,0]
Yb1 = [0,0]
Yb2 = [0,0]
Yb3 = [0,0]
MAX_L = [0,0]
CPTY = [{},{}]  # capacity[(linia,okres)]
BRAKI = []

WB = openpyxl.load_workbook('AMX_SOP_ver5.xlsx', data_only=True, read_only = True)

NBR_OF_SHIFTS, CAPACITY, NBR_WORKING_DAYS, PATH[0], PATH[1], PATH_EFF, NBR_HOURS, PATH_TPS, \
    HORIZON_W, HORIZON_D, PACKING_LINES = load_parameters(WB)
TM_LIMIT_W, TM_LIMIT_D, BACKLOGS_W, BACKLOGS_D = load_options(WB)
EFFICIENCY, LINES, GROUP, LINE_PRIORITY, SHORT_TEXT = load_efficiency(PATH_EFF)
                                                            # slownik postaci  {(materiał,linia): wydajność [kg/8h]}
                                                            #!!! uwaga wydajność w kg/8h

print(HORIZON_D,HORIZON_W)
ZAKRES[0] = [int(HORIZON_W[0].strftime('%W')), int(HORIZON_W[0].strftime('%W')) + (HORIZON_W[1] - HORIZON_W[0]).days//7]
ZAKRES[1] = [int(HORIZON_D[0].strftime('%j')), int(HORIZON_D[0].strftime('%j')) + (HORIZON_D[1] - HORIZON_D[0]).days]
DAY_DATE, WEEK_DATE = edit_date_format(ZAKRES)


print('-------------------------------- źródła danych --------------------------------------------')
print('ZLECENIA TYGODNIOWO: ', PATH[0])
print('ZLECENIA NA DZIEN: ', PATH[1])
print('WYDAJNOŚCI LINII PAKUJĄCYCH: ', PATH_EFF)
print('CZASY TPS: ',PATH_TPS)
print('-----------------------------------------------------------------------------------------')
print('--------------------------------   parametry   --------------------------------------------')
print('LICZBA ZMIAN: ', NBR_OF_SHIFTS)
print('LICZBA LINII PRACUJĄCYCH JEDNOCZEŚNIE: ', CAPACITY)
print('LICZBA DNI ROBOCZYCH: ', NBR_WORKING_DAYS)
print('LICZBA GODZIN NA ZMIANIE: ', NBR_HOURS)
print('HORYZONT PLANOWANIA (model tygodniowy): ', ZAKRES[0])
print('HORYZONT PLANOWANIA (model dzienny): ', ZAKRES[1])
print('LINIE PLASTROWNI: ', PACKING_LINES)

print('----------------------------------------------------------------------------------------')
print('--------------------------------   opcje   --------------------------------------------')
print('CZAS PRACY SOLVERA (model tygodniowy) [s]: ',TM_LIMIT_W)
print('CZAS PRACY SOLVERA (model dzienny) [s]: ',TM_LIMIT_D)
print('POZWALAJ NA ZALEGLOŚCI (model tygodniowy): ',BACKLOGS_W)
print('POZWALAJ NA ZALEGLOŚCI (model dzienny): ',BACKLOGS_D)

print('----------------------------------------------------------------------------------------')

CZASY_TPS = load_data_from_excel(PATH_TPS,'Arkusz1')


UTIL_RATE = NBR_HOURS/8                                                         # średni OEE (czas)
MAX_L[0] = [CAPACITY * NBR_WORKING_DAYS * NBR_OF_SHIFTS] * 1000           # łączna zdolność produkcyjna linii pakujących
MAX_L[1] = [CAPACITY * NBR_OF_SHIFTS] * 1000

load_exceptions_calendar(WB,HORIZON_W, HORIZON_D)
load_exceptions_lines(WB,HORIZON_W, HORIZON_D)

COEFF_X = create_coeff_x(LINE_PRIORITY)

#---------------podłączenie do serwera-----------------------
with    gp.Env(empty=True) as env:
    
        env.setParam('CSManager', 'http://10.101.5.80:61080')
        env.setParam('CSAPIACCESSID', '35aa618d-b8e1-48d7-81c9-588b771991bd')
        env.setParam('CSAPISECRET', '6fd7cb7d-567f-47ad-9868-2f88d4a7fd10')
        env.start()


        for iter in range(2):
            SALES[iter] = load_data_from_excel(PATH[iter], 'Właściwy1')
            SALES_ORDERS[iter], PRODUCTS[iter] = edit_sales_orders(SALES[iter], iter)
                                                                                                    # slownik postaci  {(materiał,tydzień): waga [kg]}
            if iter == 0:                                                                           # 1 tydzien kolejnego roku bedzie mial numer 53
                TPS = edit_tps(CZASY_TPS, PRODUCTS[0],7)

            print('ZAKRES: ', ZAKRES[iter])
            print()

            BRAKI.extend(check_production_version(PRODUCTS[iter], LINES, EFFICIENCY))
            save_braki()

            M[iter] = gp.Model('AMX_week'+str(iter))

            X[iter], S[iter], Z[iter], Y[iter], Yb1[iter], Yb2[iter], Yb3[iter] = \
                                create_variables(M[iter],ZAKRES[iter],PRODUCTS[iter],LINES,EFFICIENCY)

            if iter == 1:
                NBR_WORKING_DAYS = 1
                ERA, ERR, F, G, XX, GX = create_balance_cns2(M[iter], Yb1[iter], Yb2[iter], Yb3[iter], X[iter], Y[iter], Z[iter], S[iter], ZAKRES[iter], \
                                            PRODUCTS[iter], LINES, PACKING_LINES, SALES_ORDERS[iter], EFFICIENCY, UTIL_RATE,\
                                            MAX_L[iter], NBR_OF_SHIFTS, CPTY[iter], TPS, GROUP)


                c = [1e0, 1e2, 1e7, 1e9, 1e11, 1e6, 1e5, 1e1, 1e4, 1e1]

                M[iter].setObjective(sum(c[0] * S[iter][j] for j in S[iter]) + sum(c[1] * Z[iter][j] for j in Z[iter]) + \
                                     sum(c[2] * ERA[j] for j in ERA) + sum(c[3] * Yb1[iter][j] for j in Yb1[iter]) + \
                                     sum(c[4] * Yb2[iter][j] for j in Yb2[iter]) + sum(c[5] * Yb3[iter][j] for j in Yb3[iter]) + \
                                     sum(c[6] * G[j] for j in G) + \
                                     sum(c[7] * XX[j] for j in XX) - \
                                     sum(c[8] * COEFF_X.get((j[0],j[1]),0) * X[iter][j] for j in X[iter] if j[1] in PACKING_LINES) + \
                                     sum(c[9] * GX[j] for j in GX),
                                     GRB.MINIMIZE)

                M[iter].update()

                for licz, var in enumerate(M[iter].getVars()):
                    var.setAttr("VarName", f'var1[{licz}]')

                M[iter].Params.TimeLimit = TM_LIMIT_D
                M[iter].optimize()
                prod_total = round(sum(X[iter][j].X for j in X[iter] if len(j) == 2))
                zapas_total = round(sum(S[iter][j].X for j in S[iter]))
                zal_total = round(sum(Z[iter][j].X for j in Z[iter]))
                dobra_linia = sum(X[iter][(prod,l,t)].X for prod in PRODUCTS[iter] for l in LINES for t in range(ZAKRES[iter][0],ZAKRES[iter][1]+1) \
                                if (EFFICIENCY[(prod,l)] not in [None,0] and LINE_PRIORITY[prod] == l))
                zla_linia = sum(X[iter][(prod,l,t)].X for prod in PRODUCTS[iter] for l in LINES for t in range(ZAKRES[iter][0],ZAKRES[iter][1]+1) \
                                if (EFFICIENCY[(prod,l)] not in [None,0] and LINE_PRIORITY[prod] != l and LINE_PRIORITY[prod] != 'brak'))

                print()
                print('-------------------- PODSUMOWANIE (model dzienny) ----------------------')
                print()

                print('produkcja calkowita: ', prod_total)
                print('produkcja przedwczesna: ',zapas_total, ' = ', zapas_total/prod_total*1e5//1e2/10, '%')
                print('zaleglosc: ', zal_total, ' = ', zal_total/prod_total*1e5//1e2/10, '%')
                print('błędy pełnych zmian: ', sum(ERA[j].X for j in ERA))
                print('godziny nadliczbowe na liniach (suma): ', sum(Yb1[iter][j].X for j in Yb1[iter])*1e4//1e2/1e2)
                print('przekroczenie dziennych mocy produkcyjnych (max linii): ', sum(Yb2[iter][j].X for j in Yb2[iter]))
                print('przekroczenie mocy plastrowni (max dzienny): ', sum(Yb3[iter][j].X for j in Yb3[iter]))
                print('błędy grupowania: ', sum(GX[j].X for j in GX), ' = ',  sum(GX[j].X for j in GX)/len(GX)*1e5//1e2/10, '%')
                print('błędy linii priorytetowych: ', round(zla_linia), ' = ', zla_linia/(zla_linia + dobra_linia)*1e5//1e2/10, '%')
                print('--------------------------------------------------------------')
            else:
                create_balance_cns(M[iter], Yb1[iter], Yb2[iter], Yb3[iter], X[iter], Y[iter], Z[iter], S[iter],
                                    ZAKRES[iter], PRODUCTS[iter], LINES, PACKING_LINES, SALES_ORDERS[iter], EFFICIENCY, UTIL_RATE, \
                                    MAX_L[iter], NBR_OF_SHIFTS, NBR_WORKING_DAYS, CPTY[iter] )

                c = [1, 1e6, 1e8, 1e4, 1e1, 1e2]

                M[iter].setObjective(sum(c[0] * S[iter][j] for j in S[iter]) + sum(c[1] * Yb1[iter][j] for j in Yb1[iter]) + \
                                     sum(c[2] * Yb2[iter][j] for j in Yb2[iter]) + sum(c[3] * Yb3[iter][j] for j in Yb3[iter]) + \
                                     sum(c[4] * Z[iter][j] for j in Z[iter]) - \
                                     sum(c[5] * COEFF_X.get((j[0], j[1]), 0) * X[iter][j] for j in X[iter] if j[1] in PACKING_LINES),
                                     GRB.MINIMIZE)

                M[iter].update()

                for licz, var in enumerate( M[iter].getVars() ):
                    var.setAttr("VarName", f'var0[{licz}]')

                M[iter].Params.TimeLimit = TM_LIMIT_W
                M[iter].optimize()

                prod_total = round(sum(X[iter][j].X for j in X[iter] if len(j) == 2))
                zapas_total = round(sum(S[iter][j].X for j in S[iter]))
                zal_total = round(sum(Z[iter][j].X for j in Z[iter]))
                dobra_linia = sum(X[iter][(prod, l, t)].X for prod in PRODUCTS[iter] for l in LINES for t in
                                  range(ZAKRES[iter][0], ZAKRES[iter][1] + 1) \
                                  if (EFFICIENCY[(prod, l)] not in [None, 0] and LINE_PRIORITY[prod] == l))
                zla_linia = sum(X[iter][(prod, l, t)].X for prod in PRODUCTS[iter] for l in LINES for t in
                                range(ZAKRES[iter][0], ZAKRES[iter][1] + 1) \
                                if (EFFICIENCY[(prod, l)] not in [None, 0] and LINE_PRIORITY[prod] != l and LINE_PRIORITY[prod] != 'brak'))

                print()
                print('------------------- PODSUMOWANIE (model tygodniowy) -----------------------------------')
                print()

                print('produkcja calkowita: ', prod_total)
                print('produkcja przedwczesna: ', zapas_total, ' = ', zapas_total / prod_total * 1e5 // 1e2 / 10, '%')
                print('zaleglosc: ', zal_total, ' = ', zal_total / prod_total * 1e5 // 1e2 / 10, '%')
                print('godziny nadliczbowe na liniach (suma): ', sum(Yb1[iter][j].X for j in Yb1[iter]) * 1e4 // 1e2 / 1e2)
                print('godziny nadliczbowe na liniach (max): ', sum(Yb2[iter][j].X for j in Yb2[iter])* 1e4 // 1e2 / 1e2)
                print('przekroczenie mocy plastrowni (max tygodniowy): ', sum(Yb3[iter][j].X for j in Yb3[iter])* 1e4 // 1e2 / 1e2)
                print('błędy linii priorytetowych: ', round(zla_linia), ' = ',
                      zla_linia / (zla_linia + dobra_linia) * 1e5 // 1e2 / 10, '%')
                print('-----------------------------------------------------------------------------------------')

            try:
                print(M[iter].ObjVal)
            except AttributeError:
                break


            # dalsza czesc to zapisanie wynikow do plikow csv
            rows = []
            for prod in PRODUCTS[iter]:
                for l in LINES:
                    for t in range(ZAKRES[iter][0],ZAKRES[iter][1]+1):
                        if EFFICIENCY[(prod,l)] > 0 and X[iter][(prod,l,t)].X > 0.0000001:
                            if iter == 1:
                                o = str(DAY_DATE[t])[0:10]
                            else:
                                o = str(t)
                            if Z[iter][(prod,t)].X > 0.1:
                                znak_z = 1
                            else:
                                znak_z = 0
                            x_ratio = X[iter][(prod,l,t)].X/EFFICIENCY[(prod,l)]
                            rows.append([prod + ' (' + SHORT_TEXT[prod] + ')',
                                         l, o, X[iter][(prod, l, t)].X, 'kg', GROUP.get(prod,'brak'), LINE_PRIORITY.get(prod,'brak'),znak_z])
                            rows.append([prod + ' (' + SHORT_TEXT[prod] + ')',
                                         l, o, x_ratio * 8, 'h',  GROUP.get(prod,'brak'), LINE_PRIORITY.get(prod,'brak'),znak_z])
                            rows.append([prod + ' (' + SHORT_TEXT[prod] + ')',
                                         l, o, x_ratio / UTIL_RATE, 'zmiana', GROUP.get(prod,'brak'), LINE_PRIORITY.get(prod,'brak'),znak_z ])


            for t in range(ZAKRES[iter][0], ZAKRES[iter][1] + 1):
                suma_plastry_t = 0
                suma_plastry_t_r = 0
                suma_plastry_t += sum(X[iter][(prod,l,t)].X for prod in PRODUCTS[iter] for l in PACKING_LINES if EFFICIENCY[(prod,l)]>0)
                suma_plastry_t_r += sum(X[iter][(prod, l, t)].X/EFFICIENCY[(prod,l)] for prod in PRODUCTS[iter] \
                                     for l in PACKING_LINES if EFFICIENCY[(prod, l)] > 0)
                if iter == 1:
                    o = str(DAY_DATE[t])[0:10]
                else:
                    o = str(t)
                rows.append([0, 'plastry', o, suma_plastry_t, 'kg'])
                rows.append([0, 'plastry', o, suma_plastry_t_r * 8, 'h'])
                rows.append([0, 'plastry', o, suma_plastry_t_r/UTIL_RATE, 'zmiana'])

            with open('erp_AMX_result' + str(iter) + '.csv', 'w', encoding = 'utf-8') as f:
                # using csv.writer method from CSV package
                write = csv.writer(f)
                write.writerow(['produkt', 'linia', 'okres', 'produkcja', 'jdn', 'grupa produktowa','linia priorytetowa','zalegle'])
                write.writerows(rows)

            rows = []
            for prod in PRODUCTS[iter]:

                z = sum(Z[iter][(prod, t)].X for t in range(ZAKRES[iter][0], ZAKRES[iter][1] + 1))
                s = sum(S[iter][(prod, t)].X for t in range(ZAKRES[iter][0], ZAKRES[iter][1] + 1))

                if z > 0.01:
                    znak_z = 1
                else:
                    znak_z = 0
                if s > 0.01:
                    znak_s = 1
                else:
                    znak_s = 0

                n=0
                suma_eff=0
                for l in LINES:
                    if EFFICIENCY[(prod,l)] > 0:
                        n += 1
                        suma_eff += EFFICIENCY[(prod,l)]
                av_eff = suma_eff/n

                for t in range(ZAKRES[iter][0],ZAKRES[iter][1]+1):
                    try:
                        x = sum(X[iter][(prod, l, t)].X for l in PACKING_LINES if EFFICIENCY.get((prod,l),0) not in [None,0])
                    except KeyError:
                        x = 0

                    try:
                        s = S[iter][(prod, t)].X
                    except KeyError:
                        s = 0
                    try:
                        z = Z[iter][(prod, t)].X
                    except KeyError:
                        z = 0
                    if iter == 1:
                        o = str(DAY_DATE[t])[0:10]
                    else:
                        o = WEEK_DATE[t]
                    d = SALES_ORDERS[iter].get((prod, t), 0)


                    t_ps = TPS.get(prod,-1)
                    rows.append(([prod + ' (' + SHORT_TEXT[prod] + ')', t_ps, o, x, d, s, z, 'kg', znak_s, znak_z])) #zmiana 28.11
                    rows.append(([prod + ' (' + SHORT_TEXT[prod] + ')', t_ps, o, x * 8/ av_eff, d * 8/ av_eff, s * 8 / #zmiana 28.11
                                  av_eff, z * 8 / av_eff, 'h',znak_s, znak_z]))
                    rows.append(([prod + ' (' + SHORT_TEXT[prod] + ')', t_ps, o, x/av_eff, d/av_eff, s/av_eff, #zmiana 28.11
                                  z/av_eff, 'zmiana',znak_s, znak_z]))

                    for l in PACKING_LINES:
                        if EFFICIENCY.get((prod, l), 0) not in [None, 0]:
                            rows.append(([0, t_ps, o, x, d, s, z, 'kg', znak_s, znak_z]))
                            rows.append(([0, t_ps, o, x * 8 / av_eff, d * 8 / av_eff, s * 8 / av_eff, z * 8 / av_eff, 'h',
                                          znak_s, znak_z]))
                            rows.append(
                                ([0, t_ps, o, x / av_eff, d / av_eff, s / av_eff, z / av_eff, 'zmiana', znak_s, znak_z]))
                            break

            with open('erp_AMX_result' + str(iter) + 'sz.csv', 'w', encoding = 'utf-8', errors="ignore") as f:
                # using csv.writer method from CSV package
                write = csv.writer(f)
                write.writerow(['produkt', 'tps', 'okres', 'produkcja', 'demand', 'zapas', 'zaleglosc', 'jdn', 'przedwczesne', 'zalegle'])
                write.writerows(rows)

        capacity_summary(ZAKRES[1], ZAKRES[0], X, EFFICIENCY)

env.dispose()
