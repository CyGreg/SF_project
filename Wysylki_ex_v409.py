# transExp 26-02-2024
# Ta wersja oprogramowania do rozwiazywanie problemu wysylek eksportowych zastepuje teoretyczne ogleglosci 
# między punktami na faktyczne odleglosci z danych wejsciowych
# Tworzy macierz odległosci dla wszystkich tras między klientami i magazynami;

# generuje raport w xlsx w formacie jak w konsolidatorze logistycznym ...
#####ZMIANY BORYS##########
#dodani operatorzy
#usunięte komentarze ze środka modelu
#zmiana ograniczenia o odległości na sumę 300 między klientami nie licząc wyjazdu i powrotu do magazynu
#naprawione błędy numeryczne względem 4.0 beta
#dodany wiersz w dataframe odpowiedzialny za zamówienia przekierowane do operatrów
#funkcja celu przestawiona na aktualną wiedzę o zagadnieniu (rozładunki, załadunki itd) 
# Bazuje na wersji notatnika: 4.0.8
# nowe funkcjonalności:
# porawia buga w parametrach analizy
# aktualizacja raportu dla parametrów + zmiana indeksu k na ka

import pandas as pd
import requests
import networkx as nx
import numpy as np
import matplotlib.pyplot as plt
import math
import gurobipy as gp
from gurobipy import GRB
import random
from IPython.display import display
import time 
import openpyxl
from openpyxl import load_workbook
import xlsxwriter

import sys

#import sys
#import datetime as dt
#from datetime import date

# Plik wejściowy z danymi
# Wymagany skoroszyt z danymi wejściowymi oraz arkuszami "Do_solvera" oraz " Macierz_odl"

i_file = 'Wysylki_exp_GUI_v7_9_.xlsm'
#i_file = 'Dane_ŁŚ_uzgodnione.xlsx'
o_file = "Optymalne_trasy_"
app_name = 'transExpApp_4.0.9'
report = 0

print('++++++++++++++++++++++++++++++++++++++++++++++')
print('APLIKACJA transExpAPP @ WORK')
print('++++++++++++++++++++++++++++++++++++++++++++++')
print('\n')
print('I. Dane + wyniki:')
print('----------------------------------------------')
print(' - Wersja aplikacji:', app_name)
print(' - Plik źródłowy z danymi: ', i_file)
print(' - Plik wynikowy z trasami: ', o_file, '[data][data czas].xlsx')
      
print('----------------------------------------------')


# Wczytywanie danych z pliku Excel
df = pd.read_excel(i_file, sheet_name = 'Do_solvera', engine='openpyxl')
# Parametry 
dfp = pd.read_excel(i_file,sheet_name='Interfejs')

# Wyodrębnianie nazw klientów i przetwarzanie współrzędnych geograficznych
df['lat'] = df['GEO'].apply(lambda x: float(x.split(',')[0]))
df['lon'] = df['GEO'].apply(lambda x: float(x.split(',')[1]))

# Przygotowanie słownika odległości
klienci = df['Nazwa klienta'].tolist()
wspolrzedne = list(zip(df['lat'], df['lon']))
macierz_odleglosci = {klient: {} for klient in klienci}
koordy = {row['Nazwa klienta']: [row['GEO']] for _, row in df.iterrows()}

#<-23-01-2023
# Zmienna określająca dzień, dla którego planowane są transporty
# Sczytywana ze skoroszytu wejściowego, z arkusz 'Do_solvera'
date_analysis = str(df.iloc[1, 11])[:10]
#>-23-01-2023

# Wersja kodu na openpyxl
wb = openpyxl.load_workbook(i_file)
#wb = load_workbook(i_file), read_only =True)

ar = wb["Macierz_odl"]  

#ar = wb.get_sheet_by_name('Macierz_odl')  -- obsolete function?!

# Słownik zbierający informacje o klientach i geokordynatach
# Mogłby zostać rozbudowany o numer klienta / punktu 

gps = {}

# Dict. dist:
    #   Keys:
    #       (Nazwa klienta)        --> lepiej numer klienta
    #   Values:
    #        geokordynaty

for i in range(0, len(df)):
    if df.iloc[i, 1] not in gps.keys():
        gps[df.iloc[i, 1]] = df.iloc[i, 10]

        #Drukowanie przypisań klientów do GEO
        #print(df.iloc[i, 1], ":", gps[df.iloc[i, 1]])

# Słownik odległości pomocniczy (odległość dla geokordynatów)
macierz_odl = {}
    
    # Dict. dist:
    #   Keys:
    #       (GEO1, GEO2)        
    #   Values:
    #        Odległosc w km

k=0
for r in range(2, ar.max_row+1):
    for c in range(2, ar.max_column+1):    
        k+=1
        r_name = ar.cell(r,1).value
        c_name = ar.cell(1,c).value
        macierz_odl[r_name,c_name] = round(int(ar.cell(r,c).value)/1000,0)

        # Printy weryfikayjne:
        #print("Column", c)
        #print("Row ", r)
        #print(k, r_name, c_name, macierz_odl[r_name,c_name])

# Lista możliwych punktów w sieci transportowej
# Sczytywana z arkusza 'Do_solver'

klienci = df['Nazwa klienta'].tolist()

# Słownik odległości (dla nazw klientów)
# Lepiej żeby było dla numerów klientów 
# --> do poprawienia

macierz_odleglosci = {klient: {} for klient in klienci}

z=0    
for klient1 in klienci:     #ŁŚ: po co enumerate zamiast klient1 in kienci? indeks i nie używany
    for klient2 in klienci:
        if klient1 == klient2:
            continue
        macierz_odleglosci[klient1][klient2] = macierz_odl[gps[klient1],gps[klient2]]    

        # Drukowanie macierzy
        #print("")
        #z+=1
        #print(z)
        #print("Source: ", klient1, "GEO: ", gps[klient1])
        #print("Destination: ", klient2, "GEO: ", gps[klient2])
        #print("Distance: ", macierz_odleglosci[klient1][klient2])



G = nx.DiGraph()
for source in macierz_odleglosci:
    for target in macierz_odleglosci[source]:
        G.add_edge(source, target, weight=macierz_odleglosci[source][target])

if report == 1:
    for (u, v, wt) in G.edges.data('weight'):
        print(f"({u}, {v}, {wt:.2f} km)")

pos = {klient: (lon, lat) for klient, lat, lon in zip(df['Nazwa klienta'], df['lat'], df['lon'])}

# Skalowanie współrzędnych, aby pasowały do wykresu matplotlib
lon_range = max(df['lon']) - min(df['lon'])
lat_range = max(df['lat']) - min(df['lat'])
scale = max(lon_range, lat_range)

pos = {klient: ((lon - min(df['lon'])) / scale, (lat - min(df['lat'])) / scale) for klient, (lon, lat) in pos.items()}

# Tworzenie listy magazynów na podstawie wartości w kolumnie "Ilość całkowita"

kli = df[df["Ilość całkowita"] != 0]["Nazwa klienta"].tolist()
mag = []
source_mag = []
non_source_mag = []
posrednie = ['POSREDNIEPRUSZEWO', 'MC BYTOM']

# Iteruj przez wiersze ramki danych
for index, row in df.iterrows():
    nazwa_klienta = row['Nazwa klienta']
    zrodlo = row['Źródło']
    
    mag.append(zrodlo)  # Dodaj Źródło do listy mag
    
    if nazwa_klienta != zrodlo:
        source_mag.append(zrodlo)  # Jeśli Nazwa klienta i Źródło są różne, dodaj do source_mag

# Znajdź wartości w liście mag, które nie są w source_mag
non_source_mag = list(set(mag) - set(source_mag))
source_mag = list(set(source_mag))
mag = list(set(mag))
 
    
df_filtered = df[df['Kraj'] != 'ZAKŁADY'].copy()
######################NOWE############################################
# Przekształcanie kodów dostaw na stringi przy użyciu .loc
df_filtered.loc[:, 'Dostawa'] = df_filtered['Dostawa'].astype(str)

# Tworzenie słownika zamówień
zamowienia_dict = {
    row['Dostawa']: [row['Nazwa klienta'], row['Źródło'], row['Ilość całkowita'], row['Kraj'], row['Ilość palet']]
    for _, row in df_filtered.iterrows()
}

# Wypisanie słownika, aby sprawdzić czy wszystko działa poprawnie
if report == 1:
    for key, value in zamowienia_dict.items():
        print(f"Numer dostawy: {key}, Dane: {value}")

    
zamowienia_na_klienta = {klient: 0 for klient in kli}
for zamowienie, dane in zamowienia_dict.items():
    klient = dane[0]  # Nazwa klienta
    zamowienia_na_klienta[klient] += 1            
   
        
#####################Stałe##########################################################################

pal_op_bytom = 48 #cena za przewóz jednej palety u operatora bytom
maximum = dfp['Value'][17]  #13000maksymalna pojemnosc samochodu
maxpalet = dfp['Value'][18]    #33maksymalna pojemnosc paletowa
kurs_euro = float(dfp['Value'][38]) #4.36
kg_op_niepruszewo = float(dfp['Value'][34]) #0.083 #cena za kilo u operatora na czechy
koszt_fracht_niep = float(dfp['Value'][28]) #1200 #cena za wysłanie pojedynczego frachtu do operatora z niepruszewa
koszt_fracht_bytom = float(dfp['Value'][29]) #3300 #cena za wysłanie pojedynczego frachtu do operatora z bytomia
przel_niep_kg = float(dfp['Value'][31]) #0.02 #cena za przeładunek kilograma w niepruszewie
przel_bytom_pal = float(dfp['Value'][32]) #25 #cena za przeładunek palety w bytomiu
MAX_DISTANCE_BEZP = dfp['Value'][21] #300   #maksymalny dystans szukania kolejnego punktu w dostawach bezposrednich
MAX_DISTANCE_POS = dfp['Value'][22] #300    #to samo w posrednich
wymuszenie = dfp['Value'][23] #0.8 #pułap względem maximum masy, od którego zamówienie jest wymuszone w transporcie bezpośrednim
k = dfp['Value'][16] #5 ####################dfp['Value'][16] #ilość samochodów dostępna dla kazdego magazynu
koszt_km = float(dfp['Value'][26]) #1.54  # Koszt transportu na kilometr.
min_wyp = 0.41 #Minimalne wypełnienie w stosunku do maximum {0,1}
mecom_hu = float(dfp['Value'][36]) #123
mecom_sk = float(dfp['Value'][35]) #123


################MODEL##################################################################################################

timestr = time.strftime("%Y%m%d-%H%M%S")

with gp.Env(empty=True) as env:
    env.setParam('CSAppName', app_name)
    env.setParam('CSPriority' ,5)
    env.setParam('LogFile', 'opt_tra_' +date_analysis+'('+timestr+ ').log')
    env.setParam('CSManager', 'http://10.101.1.151:61080')
    #env.setParam('CSAPIACCESSID', 'a6dec022-37e3-4534-ade9-1dddf42d9e73')
    #env.setParam('CSAPISECRET', '2272905c-15b9-4799-988e-968f2f791556')
    env.start()
    
    with gp.Model(env=env) as m:



        # Zmienne decyzyjne



        x = m.addVars(G.edges, non_source_mag, range(k), vtype=GRB.BINARY)
        y = m.addVars(G.edges, range(k), vtype=GRB.BINARY, name="y")
        z = m.addVars(G.edges, source_mag, range(k), vtype=GRB.BINARY, name="z")
        u = m.addVars(G.nodes, non_source_mag, range(k), vtype=GRB.CONTINUOUS)
        u_z = m.addVars(G.nodes, source_mag, range(k), vtype = GRB.CONTINUOUS)
        assign_x = m.addVars(zamowienia_dict.keys(), non_source_mag, range(k), vtype=GRB.BINARY, name="assign_x")
        assign_y = m.addVars(zamowienia_dict.keys(), source_mag, non_source_mag, range(k), vtype=GRB.BINARY, name="assign_y")
        assign_z = m.addVars(zamowienia_dict.keys(), source_mag, range(k), vtype=GRB.BINARY, name="assign_z")
        assign_operator = m.addVars(zamowienia_dict.keys(), posrednie, vtype=GRB.BINARY, name="assign_operator")
        

        



########FUNKCJA CELU####################################################################################################


        
        m.setObjective(
            gp.quicksum(G.edges[i, j]['weight'] * x[i, j, d, v] * koszt_km * kurs_euro for i, j in G.edges for d in non_source_mag for v in range(k)) + #koszt kilometra naszego własnego transportu
            gp.quicksum(zamowienia_dict[zamowienie][2]*assign_operator[zamowienie, 'POSREDNIEPRUSZEWO'] * kg_op_niepruszewo * kurs_euro for zamowienie in zamowienia_dict.keys())+ #koszt za kilogram niepruszewo operator
            gp.quicksum(assign_operator[zamowienie, 'POSREDNIEPRUSZEWO']*zamowienia_dict[zamowienie][2]*koszt_fracht_niep *kurs_euro/maximum for zamowienie in zamowienia_dict.keys()) + #koszt wysłania frachtów do operatora niepruszewo
            gp.quicksum(assign_operator[zamowienie, 'MC BYTOM']*zamowienia_dict[zamowienie][2]*koszt_fracht_bytom/(maximum) for zamowienie in zamowienia_dict.keys()) + #koszt wysłania frachtu do operatora bytom
            gp.quicksum(
                assign_operator[zamowienie, 'MC BYTOM'] * zamowienia_dict[zamowienie][4] * 
                (mecom_hu * (zamowienia_dict[zamowienie][3] == 'Węgry') + 
                 mecom_sk * (zamowienia_dict[zamowienie][3] == 'Słowacja')
                 ) * kurs_euro
                for zamowienie in zamowienia_dict.keys()
            ) + #koszt za paletę u operatora bytom,
            gp.quicksum(assign_y[zamowienie, d_prim, 'POSREDNIEPRUSZEWO', v]*zamowienia_dict[zamowienie][2]*przel_niep_kg*kurs_euro for zamowienie in zamowienia_dict.keys() for v in range(k) for d_prim in source_mag if not (d_prim == 'MC NIEPRUSZEWO')) + #przeładunek niepruszewo
            gp.quicksum(assign_y[zamowienie, d_prim, 'MC BYTOM', v]*zamowienia_dict[zamowienie][4]*przel_bytom_pal for zamowienie in zamowienia_dict.keys() for v in range(k) for d_prim in source_mag) + #przeładunek bytom
            
            gp.quicksum(G.edges[i, j]['weight'] * y[i, j, v] * koszt_km*kurs_euro for i, j in G.edges for v in range(k)) + #koszt przejechania km y
            gp.quicksum(G.edges[i, j]['weight'] * z[i, j, d, v] * koszt_km*kurs_euro for i, j in G.edges for d in source_mag for v in range(k)), #koszt przejechania km z
            GRB.MINIMIZE
        )
#########SEKCJA OPERATORA############################################################################################

        
        # Ograniczenie uniemożliwiające przypisanie zamówienia do operatora w POSREDNIEPRUSZEWO,
            # jeśli kraj zamówienia nie jest Republiką Czeską
        for zamowienie, dane in zamowienia_dict.items():
            kraj = dane[3]  # Kraj zamówienia
            if kraj != "Republ. Czeska":
                m.addConstr(
                    assign_operator[zamowienie, 'POSREDNIEPRUSZEWO'] == 0,
                    name=f"no_operator_assignment_if_not_Czech_{zamowienie}"
                )
                
        for zamowienie, dane in zamowienia_dict.items():
            kraj = dane[3]  # Kraj zamówienia
            if kraj not in ["Słowacja", "Węgry"]:
                m.addConstr(
                    assign_operator[zamowienie, 'MC BYTOM'] == 0,
                    name=f"no_operator_assignment_if_not_Slovakia_or_Hungary_{zamowienie}"
                    )       
                
                

        
          # Każde zamówienie musi być przypisane do tylko jednego x lub tylko jednego z
        for zamowienie in zamowienia_dict.keys():
            m.addConstr(
                gp.quicksum(assign_x[zamowienie, d, v] for d in non_source_mag for v in range(k)) +
                gp.quicksum(assign_z[zamowienie, d, v] for d in source_mag for v in range(k))
                + gp.quicksum(assign_operator[zamowienie, d] for d in posrednie) == 1,
                name=f"assign_one_route_{zamowienie}"
            )
            
        for zamowienie, dane in zamowienia_dict.items():
            d_prim = dane[1]  # Magazyn źródłowy dla danego zamówienia
            for d in posrednie:
                # Ograniczenie zapewnia, że jeśli zamówienie jest przypisane do operatora z magazynu d, 
                # to musi istnieć aktywna trasa y prowadząca do tego magazynu
                m.addConstr(
                    assign_operator[zamowienie, d] <= gp.quicksum(y[d_prim, d, v] * assign_y[zamowienie, d_prim, d, v] for v in range(k)),
                    name=f"active_y_for_assigned_operator_{zamowienie}_{d}"
                )


 #########SEKCJA PRZYPISAŃ##################################################################################################       
      
    # Stała "wielkie M", wystarczająco duża, aby być większa niż maksymalny możliwy ładunek
        eps = 0.01
        M2 = (1+eps) * maximum
        

        for d in non_source_mag:
            for v in range(k):
                # Suma mas zamówień dla pojazdu
                total_load_x = gp.quicksum(zamowienia_dict[zamowienie][2] * assign_x[zamowienie, d, v] for zamowienie in zamowienia_dict.keys())

                # Zmienna binarna wskazująca, czy pojazd jest używany
                vehicle_used_x = m.addVar(vtype=GRB.BINARY, name=f"vehicle_used_{d}_{v}")

                # Ograniczenia zapewniające, że total_load jest albo równy zero, albo większy lub równy 0.3 maksymalnej pojemności
                m.addConstr(
                    total_load_x >= 0.3* maximum - (1 - vehicle_used_x) * M2,
                    name=f"min_load_if_used_{d}_{v}"
                )
                m.addConstr(
                    total_load_x <= vehicle_used_x * M2,
                    name=f"zero_if_not_used_{d}_{v}"
                )
        
    
        # Dodanie zmiennej binarnej wskazującej na aktywność trasy x
        is_active_x = m.addVars(non_source_mag, range(k), vtype=GRB.BINARY, name="is_active_x")
        
        
        # Ograniczenie zabraniające odwiedzenia punktu przez x, jeśli nie są do niego przypisane zamówienia
        for j in kli:  # Dla każdego klienta
            for d in non_source_mag:  # Dla każdego magazynu pośredniego
                for v in range(k):  # Dla każdego pojazdu
                    # Tworzymy wyrażenie sumujące wszystkie zmienne assign_x dla danego punktu i pojazdu
                    # Jeśli suma ta wynosi 0, oznacza to, że nie są przypisane żadne zamówienia
                    m.addConstr(
                        gp.quicksum(assign_x[zamowienie, d, v] for zamowienie, data in zamowienia_dict.items() if data[0] == j) >= 
                        gp.quicksum(x[i, j, d, v] for i in G.predecessors(j)),
                        name=f"no_visit_without_order_{j}_{d}_{v}"
                    )
        for j in kli:  # Dla każdego klienta
                    for d in source_mag:  # Dla każdego magazynu pośredniego
                        for v in range(k):  # Dla każdego pojazdu
                            # Tworzymy wyrażenie sumujące wszystkie zmienne assign_x dla danego punktu i pojazdu
                            # Jeśli suma ta wynosi 0, oznacza to, że nie są przypisane żadne zamówienia
                            m.addConstr(
                                gp.quicksum(assign_z[zamowienie, d, v] for zamowienie, data in zamowienia_dict.items() if data[0] == j) >= 
                                gp.quicksum(z[i, j, d, v] for i in G.predecessors(j)),
                                name=f"no_visit_without_order_{j}_{d}_{v}"
                            )         
                    
                    
        # Ograniczenie zapewniające, że trasa x jest aktywna tylko wtedy, gdy przypisane są do niej zamówienia
        for d in non_source_mag:
            for v in range(k):
                m.addConstr(
                    gp.quicksum(assign_x[zamowienie, d, v] for zamowienie in zamowienia_dict.keys()) >= is_active_x[d, v],
                    name=f"active_x_if_assigned_{d}_{v}"
                )
                # Zapewnienie, że trasa x jest aktywna tylko jeśli przypisane są do niej zamówienia
                m.addConstr(
                    gp.quicksum(x[i, j, d, v] for i, j in G.edges) <= len(G.edges) * is_active_x[d, v],
                    name=f"x_active_only_if_assigned_{d}_{v}"
                )
 

        # Ograniczenie uniemożliwiające przypisanie zamówienia do y, jeśli jest przypisane do z
        for zamowienie in zamowienia_dict.keys():
            for d_prim in source_mag:
                for d in non_source_mag:
                    # Łączna suma przypisań zamówienia do tras y i z
                    total_assigns_y = gp.quicksum(assign_y[zamowienie, d_prim, d, v] for v in range(k))
                    total_assigns_z = gp.quicksum(assign_z[zamowienie, d_prim, v] for v in range(k))

                    # Dodanie ograniczenia
                    m.addConstr(
                        total_assigns_y + total_assigns_z <= 1,
                        name=f"no_y_if_assigned_to_z_{zamowienie}_{d_prim}_{d}"
                    )







        for zamowienie in zamowienia_dict.keys():
            for d in source_mag:
                for v in range(k):
                    m.addConstr(
                        gp.quicksum(assign_operator[zamowienie, d_prime] for d_prime in posrednie) + assign_z[zamowienie, d, v] <= 1,
                        name=f"no_z_if_assigned_to_operator_{zamowienie}_{d}_{v}"
                    )
        for zamowienie in zamowienia_dict.keys():
            for d in non_source_mag:
                for v in range(k):
                    m.addConstr(
                        gp.quicksum(assign_operator[zamowienie, d_prime] for d_prime in posrednie) + assign_x[zamowienie, d, v] <= 1,
                        name=f"no_z_if_assigned_to_operator_{zamowienie}_{d}_{v}"
                    )           



        # Ograniczenie przypisania zamówień do tras z, które startują z magazynu źródłowego zamówienia
        for zamowienie, dane in zamowienia_dict.items():
            d_prim = dane[1]  # Magazyn źródłowy dla danego zamówienia
            for d in source_mag:
                for v in range(k):
                    if d != d_prim:
                        m.addConstr(
                            assign_z[zamowienie, d, v] == 0,
                            name=f"restrict_z_not_from_source_{zamowienie}_{d}_{v}"
                        )





        for zamowienie, dane in zamowienia_dict.items():
            klient = dane[0]  # Końcowa destynacja zamówienia
            d_prim = dane[1]  # Magazyn źródłowy zamówienia

            for d in posrednie:
                for v in range(k):
                    # Ograniczenie zapewnia, że jeśli zamówienie jest przypisane do trasy y z magazynu d_prim do magazynu d,
                    # to musi istnieć aktywna trasa x z magazynu d do klienta LUB zamówienie musi być przypisane do operatora z tego magazynu
                    m.addConstr(
                        y[d_prim, d, v] * assign_y[zamowienie, d_prim, d, v] <= gp.quicksum(x[j, klient, d, v_prime] * assign_x[zamowienie, d, v_prime] for j in G.predecessors(klient) for v_prime in range(k)) + assign_operator[zamowienie, d],
                        name=f"y_active_and_x_or_operator_destination_{zamowienie}_{d_prim}_{d}_{v}"
                    )



        # Jeśli zamówienie jest przypisane do x, to musi być przypisane do y
        for zamowienie, dane in zamowienia_dict.items():
            d_prim = dane[1]# Magazyn źródłowy dla danego zamówienia
            for d in non_source_mag:
                for v in range(k):
                    m.addConstr(
                        assign_x[zamowienie, d, v] <= gp.quicksum(y[d_prim, d, v_prime] * assign_y[zamowienie, d_prim, d, v_prime] for v_prime in range(k)),
                        name=f"active_and_assigned_y_{zamowienie}_{d}_{v}"
                    )


        # Zamówienia mogą być przypisane tylko do aktywnych tras
        for zamowienie, dane in zamowienia_dict.items():
            klient = dane[0]
            for d in non_source_mag:
                for v in range(k):
                    m.addConstr(assign_x[zamowienie, d, v] <= gp.quicksum(x[i, klient, d, v] for i in G.predecessors(klient)),
                                name=f"active_route_x_{zamowienie}_{d}_{v}")
            for d in source_mag:
                for v in range(k):
                    m.addConstr(assign_z[zamowienie, d, v] <= gp.quicksum(z[i, klient, d, v] for i in G.predecessors(klient)),
                                name=f"active_route_z_{zamowienie}_{d}_{v}")
            for d in non_source_mag:
                for v in range(k):
                    m.addConstr(assign_y[zamowienie, d_prim, d, v] <= gp.quicksum(y[d_prim, i, v] for i in G.successors(d_prim)),
                                name=f"active_route_y_{zamowienie}_{d_prim}_{d}_{v}")


        # Ograniczenie pojemności pojazdu
        for v in range(k):
            for d in non_source_mag:
                m.addConstr(
                    gp.quicksum(zamowienia_dict[zamowienie][2] * assign_x[zamowienie, d, v] for zamowienie in zamowienia_dict.keys()) <= maximum,
                    name=f"capacity_x_vehicle_{d}_{v}"
                )

                m.addConstr(
                    gp.quicksum(zamowienia_dict[zamowienie][4] * assign_x[zamowienie, d, v] for zamowienie in zamowienia_dict.keys()) <= maxpalet,
                    name=f"capacity_x_vehicle_{d}_{v}"
                )
            for d in source_mag:
                for d_prime in non_source_mag:
                    m.addConstr(
                        gp.quicksum(zamowienia_dict[zamowienie][2] * assign_y[zamowienie, d, d_prime, v] for zamowienie in zamowienia_dict.keys()) <= maximum,
                        name=f"capacity_y_vehicle_{d}_{d_prime}_{v}"
                    )
                    m.addConstr(
                        gp.quicksum(zamowienia_dict[zamowienie][4] * assign_y[zamowienie, d, d_prime, v] for zamowienie in zamowienia_dict.keys()) <= maxpalet,
                        name=f"capacity_y_vehicle_{d}_{d_prime}_{v}"
                    )
            for d in source_mag:
                m.addConstr(
                    gp.quicksum(zamowienia_dict[zamowienie][2] * assign_z[zamowienie, d, v] for zamowienie in zamowienia_dict.keys()) <= maximum,
                    name=f"capacity_z_vehicle_{d}_{v}"
                )
                m.addConstr(
                    gp.quicksum(zamowienia_dict[zamowienie][4] * assign_z[zamowienie, d, v] for zamowienie in zamowienia_dict.keys()) <= maxpalet,
                    name=f"capacity_z_vehicle_{d}_{v}"
                )    



        double_mag = "MC NIEPRUSZEWO"   #jezeli w przyszlosci pojawią sie magazyny o tym samym charakterze, zrobię automatyzację
        #póki co mamy ograniczony zestaw danych więc jest zrobione tak

        for zamowienie, dane in zamowienia_dict.items():
            klient = dane[0]  # Nazwa klienta
            masa_zamowienia = dane[2]  # Masa zamówienia
            d_prim = dane[1]  # Magazyn źródłowy dla zamówienia

            # Sprawdzenie, czy magazyn źródłowy nie jest MC NIEPRUSZEWO
            if d_prim != double_mag:
                # Jeśli magazyn źródłowy nie jest MC NIEPRUSZEWO i masa zamówienia wymaga dostawy bezpośredniej
                if masa_zamowienia >= wymuszenie * maximum:
                    m.addConstr(
                        gp.quicksum(assign_z[zamowienie, d_prim, v] for v in range(k)) == 1,
                        name=f"direct_delivery_{zamowienie}"
                    )

                    # Zapewnienie, że to zamówienie nie jest dostarczane przez x
                    for d in non_source_mag:
                        for v in range(k):
                            m.addConstr(
                                assign_x[zamowienie, d, v] == 0,
                                name=f"no_indirect_delivery_{zamowienie}_{d}_{v}"
                            )
                    for d_prime in posrednie:
                        m.addConstr(assign_operator[zamowienie, d_prime] == 0,
                                   name=f"no_operator_delivery_{zamowienie}_{d_prime}")
            else:
                # Dla MC NIEPRUSZEWO nie wymuszamy dostaw bezpośrednich
                for v in range(k):
                    m.addConstr(
                        assign_z[zamowienie, d_prim, v] == 0,
                        name=f"optional_direct_delivery_{zamowienie}_{d_prim}_{v}"
                    )
                    # Można również zezwolić na przypisanie zamówienia do x w przypadku MC NIEPRUSZEWO
                    for d in non_source_mag:
                        m.addConstr(
                            assign_x[zamowienie, d, v] <= 1,
                            name=f"optional_indirect_delivery_{zamowienie}_{d}_{v}"
                        )


        # Ograniczenie zabraniające wysyłek bezpośrednich z MC NIEPRUSZEWO
        for zamowienie, dane in zamowienia_dict.items():
            d_prim = dane[1]  # Magazyn źródłowy dla danego zamówienia

            # Sprawdzenie, czy magazyn źródłowy to MC NIEPRUSZEWO
            if d_prim == double_mag:
                for v in range(k):
                    m.addConstr(
                        gp.quicksum(assign_z[zamowienie, d_prim, v] for zamowienie in zamowienia_dict.keys()) == 0,
                        name=f"no_direct_delivery_from_mc_niepruszewo_{v}"
                    )                


        for zamowienie, dane in zamowienia_dict.items():
            klient = dane[0]  # Nazwa klienta
            kraj = dane[3]  # Kraj klienta

            # Dla klientów z Węgier i Słowacji
            if kraj in ["Węgry", "Słowacja"]:
                for d in non_source_mag:
                    if d != "MC BYTOM":
                        for v in range(k):
                            m.addConstr(
                                assign_x[zamowienie, d, v] == 0,
                                name=f"restrict_order_{zamowienie}_from_{d}_to_HU_SK"
                            )

            # Dla klientów z Republiki Czeskiej
            if kraj == "Republ. Czeska":
                for d in non_source_mag:
                    if d != "POSREDNIEPRUSZEWO":
                        for v in range(k):
                            m.addConstr(
                                assign_x[zamowienie, d, v] == 0,
                                name=f"restrict_order_{zamowienie}_from_{d}_to_Czech"
                            )



        # Stała "wielkie M", wystarczająco duża, aby być większa niż maksymalny możliwy ładunek
        M = (1+eps) * maximum

        for d in source_mag:
            for v in range(k):
                # Suma mas zamówień dla pojazdu
                total_load = gp.quicksum(zamowienia_dict[zamowienie][2] * assign_z[zamowienie, d, v] for zamowienie in zamowienia_dict.keys())

                # Zmienna binarna wskazująca, czy pojazd jest używany
                vehicle_used = m.addVar(vtype=GRB.BINARY, name=f"vehicle_used_{d}_{v}")

                # Ograniczenia zapewniające, że total_load jest albo równy zero, albo większy lub równy 0.3 maksymalnej pojemności
                m.addConstr(
                    total_load >= min_wyp * maximum - (1 - vehicle_used) * M,
                    name=f"min_load_if_used_{d}_{v}"
                )
                m.addConstr(
                    total_load <= vehicle_used * M,
                    name=f"zero_if_not_used_{d}_{v}"
                )

        ###technika duzego M została zastosowana w celu mozliwosci utworzenia dolnego limitu załadunku samochodu,
        ## mozna to zmieniac z poziomu interfejsu zgodnie z wymaganiami


        for zamowienie in zamowienia_dict.keys():
            # Ograniczenie dla trasy x
            m.addConstr(
                gp.quicksum(assign_x[zamowienie, d, v] for d in non_source_mag for v in range(k)) <= 1,
                name=f"assign_x_limit_{zamowienie}"
            )

            # Ograniczenie dla trasy y
            m.addConstr(
                gp.quicksum(assign_y[zamowienie, d_prim, d, v] for d_prim in source_mag for d in non_source_mag for v in range(k)) <= 1,
                name=f"assign_y_limit_{zamowienie}"
            )

            # Ograniczenie dla trasy z
            m.addConstr(
                gp.quicksum(assign_z[zamowienie, d, v] for d in source_mag for v in range(k)) <= 1,
                name=f"assign_z_limit_{zamowienie}"
            )
            
            
            m.addConstr(gp.quicksum(assign_operator[zamowienie,d] for d in posrednie) <= 1)
            


#########SEKCJA ZMIENNEJ Z###########################################################################

        for j in kli:
            for d in source_mag:
                for v in range(k):
                    m.addConstr(gp.quicksum(z[i, j, d, v] for i in G.predecessors(j)) == gp.quicksum(z[j, i, d, v] for i in G.successors(j)))




        for d in source_mag:
            for v in range(k):
                m.addConstr(gp.quicksum(z[d, j, d, v] for j in G.successors(d)) == gp.quicksum(z[j, d, d, v] for j in G.predecessors(d)))



        for d in source_mag:
            for dp in mag:
                if d != dp:
                    for v in range(k):
                        m.addConstr(gp.quicksum(z[i, dp, d, v] for i in G.predecessors(dp)) == 0)


       
                
        for d in source_mag:
            for v in range(k):
                # Suma odległości między klientami dla trasy
                m.addConstr(
                    gp.quicksum(G.edges[i, j]['weight'] * z[i, j, d, v] for i, j in G.edges if i != d and j != d) <= MAX_DISTANCE_POS,
                    name=f"max_sum_distance_for_x_{d}_{v}"
                )
        
                            
                            
                            
        for d in source_mag:
            for v in range(k):
                # Ograniczenie opuszczenia magazynu - pojazd może opuścić magazyn tylko raz
                m.addConstr(
                    gp.quicksum(z[d, j, d, v] for j in G.successors(d)) <= 1,
                    name=f"leave_warehouse_once_{d}_{v}"
                )

                # Ograniczenie powrotu do magazynu - pojazd może powrócić do magazynu tylko raz
                m.addConstr(
                    gp.quicksum(z[i, d, d, v] for i in G.predecessors(d)) <= 1,
                    name=f"return_to_warehouse_once_{d}_{v}"
                )                    



        
        # Ograniczenie maksymalnej liczby krawędzi na trasie do czterech
        for d in source_mag:
            for v in range(k):
                m.addConstr(gp.quicksum(z[i, j, d, v] for i, j in G.edges) <= 4, name=f"max_four_edges_{d}_{v}")
        for d in non_source_mag:
            for v in range(k):
                m.addConstr(gp.quicksum(x[i, j, d, v] for i, j in G.edges) <= 4)



        ############################SEKCJA DOTYCZĄCA ZMIENNEJ X#######################################################################

        #Maksymalna liczba krawędzi
        for d in non_source_mag:
            for v in range(k):
                m.addConstr(gp.quicksum(x[i, j, d, v] for i, j in G.edges) <= 4)

        # Ograniczenia dotyczące punktów dostaw
        for j in kli:
            for d in non_source_mag:
                for v in range(k):
                    m.addConstr(gp.quicksum(x[i, j, d, v] for i in G.predecessors(j)) == gp.quicksum(x[j, i, d, v] for i in G.successors(j)))

        for i in kli:
            # Ograniczenie mówiące, że liczba odwiedzin klienta 'i' musi być mniejsza lub równa liczbie jego zamówień
            m.addConstr(
                gp.quicksum(z[j, i, d, v] for j in G.predecessors(i) for d in source_mag for v in range(k)) + 
                gp.quicksum(x[j, i, d, v] for j in G.predecessors(i) for d in non_source_mag for v in range(k)) <= zamowienia_na_klienta[i], 
                name=f"visit_limit_{i}"
            )



        for d in non_source_mag:
            for v in range(k):
                m.addConstr(gp.quicksum(x[d, j, d, v] for j in G.successors(d)) == gp.quicksum(x[j, d, d, v] for j in G.predecessors(d)))



       
                            
        for d in non_source_mag:
            for v in range(k):
                # Suma odległości między klientami dla trasy
                m.addConstr(
                    gp.quicksum(G.edges[i, j]['weight'] * x[i, j, d, v] for i, j in G.edges if i != d and j != d) <= MAX_DISTANCE_POS,
                    name=f"max_sum_distance_for_x_{d}_{v}"
                )

        
        
        for d in non_source_mag:
            for v in range(k):
                # Ograniczenie opuszczenia magazynu - pojazd może opuścić magazyn tylko raz
                m.addConstr(
                    gp.quicksum(x[d, j, d, v] for j in G.successors(d)) <= 1,
                    name=f"leave_warehouse_once_{d}_{v}"
                )

                # Ograniczenie powrotu do magazynu - pojazd może powrócić do magazynu tylko raz
                m.addConstr(
                    gp.quicksum(x[i, d, d, v] for i in G.predecessors(d)) <= 1,
                    name=f"return_to_warehouse_once_{d}_{v}"
                )                    


        ################OGRANICZENIA MTZ#########################################################################################

        for i in kli:
            for j in kli:
                if i != j:
                    for d in source_mag:
                        for v in range(k):
                            m.addConstr(u_z[i, d, v] - u_z[j, d, v] + len(kli) * z[i, j, d, v] <= len(kli) - 1)



        for i in kli:
            for j in kli:
                if i != j:
                    for d in non_source_mag:
                        for v in range(k):
                            m.addConstr(u[i, d, v] - u[j, d, v] + len(kli) * x[i, j, d, v] <= len(kli) - 1)

        # Ograniczenie zakazujące odwiedzania innych magazynów
        for d in non_source_mag:
            for dp in mag:
                if d != dp:
                    for v in range(k):
                        m.addConstr(gp.quicksum(x[i, dp, d, v] for i in G.predecessors(dp)) == 0)




        # Optymalizacja modelu
        m.setParam(GRB.Param.TimeLimit, 70)
        m.setParam(GRB.Param.IntegralityFocus, 1)
        m.setParam(GRB.Param.MIPFocus, 2)
        m.setParam(GRB.Param.Threads, 0)
        m.optimize()

     
        # Wizualizacja wyników
        plt.figure(figsize=(20, 20))

        # Rysowanie punktów dostaw
        nx.draw_networkx_nodes(G, pos=pos, nodelist=kli, node_color='lightblue', node_size=300)

        # Rysowanie magazynów
        nx.draw_networkx_nodes(G, pos=pos, nodelist=mag, node_color='magenta', node_size=1000)


        # Rysowanie etykiet punktów
        nx.draw_networkx_labels(G, pos=pos)

        colors = ['r', 'g', 'b', 'y', 'c']

        for v in range(k):
            for d in non_source_mag:
                route_edges = [(i, j) for i, j in G.edges if x[i, j, d, v].x > 0.5]
                nx.draw_networkx_edges(G, pos=pos, edgelist=route_edges, edge_color=colors[v % len(colors)], width=2, arrows=True)

        for v in range(k):
            route_edges_y = [(i, j) for i, j in G.edges if y[i, j, v].x > 0.5]
            nx.draw_networkx_edges(G, pos=pos, edgelist=route_edges_y, edge_color=colors[v % len(colors)], width=2, style='dashed', arrows=True)        
        


        for v in range(k):
            for d in source_mag:
                route_edges_z = [(i, j) for i, j in G.edges if z[i, j, d, v].x > 0.5]
                nx.draw_networkx_edges(G, pos=pos, edgelist=route_edges_z, edge_color=colors[v % len(colors)], width=2, arrows=True, style = 'dashdot')

        plt.title("Wizualizacja optymalnych tras_" + date_analysis + '(' + timestr + ')')
        
        plt.savefig(o_file+ "_"+date_analysis+'('+timestr+').png')
        #plt.show()        
        #plt.close()

        def trace_route(start_point, vehicle, route_var, G, storage):
                current_point = start_point
                route = [current_point]

                while True:
                    next_point_candidates = []
                    for i, j in G.edges:
                        var = route_var.get((current_point, j, storage, vehicle))
                        if isinstance(var, gp.Var) and var.X > 0.5:
                            next_point_candidates.append(j)

                    if not next_point_candidates or (next_point_candidates[0] == start_point):
                        if current_point != start_point: 
                            route.append(start_point)
                        break

                    next_point = next_point_candidates[0]
                    route.append(next_point)
                    current_point = next_point

                return route


        df_trasy_x = []
        df_trasy_y = []
        df_trasy_z = []

        # Przetwarzanie tras x
        for d in non_source_mag:
            for v in range(k):
                assigned_orders = [zamowienie for zamowienie in zamowienia_dict.keys() if assign_x[zamowienie, d, v].X > 0.5]
                if assigned_orders:
                    route = trace_route(d, v, x, G, d)
                    total_mass = sum(zamowienia_dict[zamowienie][2] for zamowienie in assigned_orders)
                    total_pallets = sum(int(zamowienia_dict[zamowienie][4]) for zamowienie in assigned_orders)
                     # Obliczanie całkowitej długości trasy bez magazynu
                    total_distance = sum(G.edges[route[i], route[i + 1]]['weight'] for i in range(len(route)-1))

                    cost_distance = total_distance * koszt_km * kurs_euro

                    
                    additional_cost = total_pallets * przel_bytom_pal if d == 'MC BYTOM' else total_mass * przel_niep_kg * kurs_euro
                    df_trasy_x.append({
                        'Magazyn Pośredni': d,
                        'Numer Pojazdu': v,
                        'Trasa': route,
                        'Całkowita Masa': total_mass,
                        'Ilość palet': total_pallets,
                        'Ilość w kg': total_mass,
                        'Zamówienia': assigned_orders,
                        'Źródło dostawy': zamowienia_dict[assigned_orders[0]][1],
                        'Koszt':  cost_distance + additional_cost
                    })

        # Przetwarzanie tras y
        for d_prim in source_mag:
            for d in non_source_mag:
                for v in range(k):
                    assigned_orders = [zamowienie for zamowienie in zamowienia_dict.keys() if assign_y[zamowienie, d_prim, d, v].X > 0.5]
                    if assigned_orders:
                        total_mass = sum(zamowienia_dict[zamowienie][2] for zamowienie in assigned_orders)
                        total_pallets = sum(int(zamowienia_dict[zamowienie][4]) for zamowienie in assigned_orders)
                        distance = G.edges[d_prim, d]['weight']
                        df_trasy_y.append({
                            'Numer Trasy': v,
                            'Magazyn Źródłowy': d_prim,
                            'Punkt na trasie': 0,
                            'Nazwa punktu': d,
                            'Zamówienia' : assigned_orders,
                            'Przejechane KM': macierz_odleglosci[d_prim][d],
                            'Ilość w kg': total_mass,
                            'Ilość w paletach': total_pallets,
                            'GEOKOORDYNATY' : koordy[d][2:22],
                            'Źródło dostawy': zamowienia_dict[assigned_orders[0]][1],
                            'Koszt': distance * koszt_km * kurs_euro
                            
                        })

        # Przetwarzanie tras z
        for d in source_mag:
            for v in range(k):
                assigned_orders = [zamowienie for zamowienie in zamowienia_dict.keys() if assign_z[zamowienie, d, v].X > 0.5]
                
                
                if assigned_orders:
                    route = trace_route(d, v, z, G, d)
                    total_mass = sum(zamowienia_dict[zamowienie][2] for zamowienie in assigned_orders)
                    total_pallets = sum(int(zamowienia_dict[zamowienie][4]) for zamowienie in assigned_orders)
                    total_distance = sum(G.edges[route[i], route[i + 1]]['weight'] for i in range(len(route)-1))

                    df_trasy_z.append({
                        'Magazyn Źródłowy': d,
                        'Numer Pojazdu': v,
                        'Trasa': route,
                        'Całkowita Masa': total_mass,
                        'Ilość palet': total_pallets,
                        'Zamówienia' : assigned_orders,
                        'Źródło dostawy': zamowienia_dict[assigned_orders[0]][1],
                        'Koszt': total_distance * koszt_km * kurs_euro
                    })

        # Konwersja list na DataFrame
        df_trasy_x = pd.DataFrame(df_trasy_x)
        df_trasy_y = pd.DataFrame(df_trasy_y)
        df_trasy_z = pd.DataFrame(df_trasy_z)

        display(df_trasy_x)
        display(df_trasy_z)
        display(df_trasy_y)
        
###############################################################Nowe Dataframy#################################################        

        report_mod = 0
        
        new_rows_y = []
        num_trasy = 0
        for index, row in df_trasy_y.iterrows():
            
            magazyn_zrodlowy = row['Magazyn Źródłowy']
            magazyn_docelowy = row['Nazwa punktu']
            zamówienia = row['Zamówienia']
            numer_trasy = row['Numer Trasy']
            masa = row['Ilość w kg']
            palety = row['Ilość w paletach']
            source = row['Źródło dostawy']
            koszt = row['Koszt']

            v = index  # zakładam, że index jest równoznaczny z numerem pojazdu

            koord_list = []
            koord = str(koordy[magazyn_zrodlowy])
            koord_list = koord.split(',')
            lati = koord_list[0].replace("['","")
            long = koord_list[1].replace("']","")

            if magazyn_zrodlowy == "MC NIEPRUSZEWO" and magazyn_docelowy == "POSREDNIEPRUSZEWO":
            #if magazyn_zrodlowy == "SZCZECIN" and magazyn_docelowy == "MC BYTOM":
                continue    
            else:
                num_trasy +=1
                # Dodajemy punkt zerowy
                if report_mod ==1:
                    new_rows_y.append({
                        'Nr. Trasy': num_trasy,  #numer_trasy, 
                        'Źródło dostawy': source,
                        'Źródło': magazyn_zrodlowy,
                        'Punkt na trasie': 0,
                        'Nazwa punktu': magazyn_zrodlowy,
                        'Zamówienia': zamówienia,
                        'Przejechane KM': 0,
                        'Ilość w kg': masa,
                        'Ilość w paletach': palety,
                        'GEOOKOORDYNATY': koord[2:22],
                        'Szerokość': lati,
                        'Długość': long,
                        'Data przejazdu' : date_analysis,
                        'Koszt': None,
                        'Typ trasy': 'ZRODLO-MAGAZYN',
                        'Koszt':koszt
                        
                        
                        
                    })
    
                # Dodajemy punkt docelowy
                new_rows_y.append({
                    'Nr. Trasy': num_trasy,  #numer_trasy, 
                    'Źródło dostawy': source,
                    'Źródło': magazyn_zrodlowy,
                    'Punkt na trasie': 1,
                    'Nazwa punktu': magazyn_docelowy,
                    'Zamówienia': zamówienia,
                    'Przejechane KM': macierz_odleglosci[magazyn_zrodlowy][magazyn_docelowy],
                    'Ilość w kg': masa,
                    'Ilość w paletach': palety,
                    'GEOOKOORDYNATY': koord[2:22],
                    'Szerokość': lati,
                    'Długość': long,
                    'Data przejazdu' : date_analysis,
                    'Koszt': koszt,
                    'Typ trasy': 'ZRODLO-MAGAZYN'
                    
                })
    
                if report_mod ==1:
                    # Dodajemy punkt powrotny do magazynu źródłowego
                    new_rows_y.append({
                        'Nr. Trasy': num_trasy,  #numer_trasy, 
                        'Źródło dostawy': source,
                        'Źródło': magazyn_zrodlowy,
                        'Punkt na trasie': 2,
                        'Nazwa punktu': magazyn_zrodlowy,
                        'Zamówienia': "",
                        'Przejechane KM': macierz_odleglosci[magazyn_docelowy][magazyn_zrodlowy],
                        'Ilość w kg': 0,  # Założenie, że powrót jest bez ładunku
                        'Ilość w paletach': 0,  # Założenie, że powrót jest bez ładunku
                        'GEOOKOORDYNATY': koord[2:22],
                        'Szerokość': lati,
                        'Długość': long,
                        'Data przejazdu' : date_analysis,
                        'Koszt': None,
                        'Typ trasy': 'ZRODLO-MAGAZYN'

                    })

        # Tworzymy DataFrame z listy new_rows_y
        df_trasy_y_expanded = pd.DataFrame(new_rows_y)
        display(df_trasy_y_expanded)
    
        new_rows_x = []
        for index, row in df_trasy_x.iterrows():
            num_trasy +=1
            magazyn = row['Magazyn Pośredni']
            trasa = row['Trasa']
            zamówienia = row['Zamówienia']
            v = row['Numer Pojazdu']
            source = row['Źródło dostawy']
            koszt = row['Koszt']
            masa = row['Ilość w kg']

            # Iterowanie przez każdy punkt na trasie
            for point_index, point in enumerate(trasa):
                zamowienia_dla_punktu = [
                    key for key in zamówienia if assign_x.get((key, magazyn, v)).X == 1 and zamowienia_dict[key][0] == point
                ] if point_index != 0 else [
                    key for key in zamówienia if assign_x.get((key, magazyn, v)).X == 1
                ]

                koord_list = []
                koord = str(koordy[point])
                koord_list = koord.split(',')
                lati = koord_list[0].replace("['","")
                long = koord_list[1].replace("']","")

                if point_index > 0 and point_index < len(trasa)-1:
                    new_row = {
                        #'Nr. Trasy': index + 1, 
                        'Nr. Trasy': num_trasy, 
                        'Źródło dostawy': source,
                        'Źródło': magazyn,
                        'Punkt na trasie': point_index, 
                        'Nazwa punktu': point, 
                        'Zamówienia': zamowienia_dla_punktu,
                        'Przejechane KM': 0 if point_index == 0 else macierz_odleglosci[trasa[point_index - 1]][point],
                        'Ilość w kg': sum(zamowienia_dict[key][2] for key in zamowienia_dla_punktu), 
                        'Ilość w paletach': sum(zamowienia_dict[key][4] for key in zamowienia_dla_punktu),  
                        'GEOOKOORDYNATY': koord[2:22], 
                        'Szerokość': lati,
                        'Długość': long,
                        'Data przejazdu' : date_analysis,
                        'Koszt': koszt * (sum(zamowienia_dict[key][2] for key in zamowienia_dla_punktu))/(masa),
                        'Typ trasy': 'MAGAZYN-KLIENT'
                    }

                    new_rows_x.append(new_row)

        
        
                # Inicjalizacja list zamówień dla obu magazynów
        
        df_trasy_x_expanded = pd.DataFrame(new_rows_x)
        # Przygotowanie nowych wierszy jako listy słowników
        new_rows = []
        
        
        assigned_orders_posredniepruszewo = [zamowienie for zamowienie, magazyn in assign_operator.keys() if assign_operator[zamowienie, magazyn].X > 0.5 and magazyn == 'POSREDNIEPRUSZEWO']
        assigned_orders_mc_bytom = [zamowienie for zamowienie, magazyn in assign_operator.keys() if assign_operator[zamowienie, magazyn].X > 0.5 and magazyn == 'MC BYTOM']

        
        if assigned_orders_posredniepruszewo:
            new_rows.append({
                
                'Nr. Trasy': df_trasy_x_expanded['Nr. Trasy'].max() + 1 if not df_trasy_x_expanded.empty else 1,
                'Źródło dostawy': None,
                'Źródło dostawy': source,
                'Źródło': 'MC NIEPRUSZEWO',
                'Punkt na trasie': 100,
                'Nazwa punktu': 'OPERATOR_CZECHY',
                'Zamówienia': assigned_orders_posredniepruszewo,
                'Przejechane KM': 490,
                'Ilość w kg': sum(zamowienia_dict[zamowienie][2] for zamowienie in assigned_orders_posredniepruszewo),
                'Ilość w paletach': sum(zamowienia_dict[zamowienie][4] for zamowienie in assigned_orders_posredniepruszewo),
                'GEOOKOORDYNATY': '49.9696553, 14.6288109',
                'Szerokość': '49.9696553',
                'Długość': '14.6288109',
                'Data przejazdu': date_analysis,
                'Koszt' :  (sum(zamowienia_dict[key][2] for key in assigned_orders_posredniepruszewo))*(koszt_fracht_niep*kurs_euro/(maximum) + przel_niep_kg) + sum(zamowienia_dict[key][2] for key in assigned_orders_posredniepruszewo) * kg_op_niepruszewo ,
                'Typ trasy': 'OPERATOR'
            })

        
        if assigned_orders_mc_bytom:
            new_rows.append({
                'Nr. Trasy': df_trasy_x_expanded['Nr. Trasy'].max() + 2 if not df_trasy_x_expanded.empty else 2,
                #'Źródło dostawy': None,
                'Źródło dostawy': source,
                'Źródło': 'MC BYTOM',
                'Punkt na trasie': 100,
                'Nazwa punktu': 'OPERATOR_HU_SK',
                'Zamówienia': assigned_orders_mc_bytom,
                'Przejechane KM': 400,
                'Ilość w kg': sum(zamowienia_dict[zamowienie][2] for zamowienie in assigned_orders_mc_bytom),
                'Ilość w paletach': sum(zamowienia_dict[zamowienie][4] for zamowienie in assigned_orders_mc_bytom),
                'GEOOKOORDYNATY': '48.4440497, 19.7019243',
                'Szerokość': '48.4440497',
                'Długość': '19.7019243',
                'Data przejazdu': date_analysis,
                'Koszt': (sum(zamowienia_dict[key][2] for key in assigned_orders_mc_bytom))*koszt_fracht_bytom/(maximum) + sum(zamowienia_dict[key][4] for key in assigned_orders_mc_bytom) * (pal_op_bytom * kurs_euro + przel_bytom_pal) ,
                'Typ trasy': 'OPERATOR'
            })

        #new_rows_x.append(new_rows)
        new_rows_x.extend(new_rows) 
        
        # Tworzenie DataFrame z nowymi wierszami
        df_new_rows = pd.DataFrame(new_rows)

        # Dodawanie nowych wierszy do df_trasy_x_expanded
        df_trasy_x_expanded = pd.concat([df_trasy_x_expanded, df_new_rows], ignore_index=True)
        #display(df_trasy_x_expanded)
        
        new_rows_z = []
        for index, row in df_trasy_z.iterrows():
            num_trasy +=1
            magazyn = row['Magazyn Źródłowy']
            trasa = row['Trasa']
            zamówienia = row['Zamówienia']
            v = row['Numer Pojazdu']
            source = row['Źródło dostawy']
            koszt = row['Koszt']
            masa = row['Całkowita Masa']

            # Iterowanie przez każdy punkt na trasie
            for point_index, point in enumerate(trasa):
                zamowienia_dla_punktu = [
                    key for key in zamówienia if assign_z.get((key, magazyn, v)).X == 1 and zamowienia_dict[key][0] == point
                ] if point_index != 0 else [
                    key for key in zamówienia if assign_z.get((key, magazyn, v)).X == 1
                ]
                koord_list = []
                koord = str(koordy[point])
                koord_list = koord.split(',')
                lati = koord_list[0].replace("['","")
                long = koord_list[1].replace("']","")
                
                if point_index > 0 and point_index < len(trasa)-1:
                    new_row_z = {
                        #'Nr. Trasy': index + 1, 
                        'Nr. Trasy': num_trasy, 
                        'Źródło dostawy': source,
                        'Źródło': magazyn,
                        'Punkt na trasie': point_index, 
                        'Nazwa punktu': point, 
                        'Zamówienia': zamowienia_dla_punktu,
                        'Przejechane KM': 0 if point_index == 0 else macierz_odleglosci[trasa[point_index - 1]][point],
                        'Ilość w kg': sum(zamowienia_dict[key][2] for key in zamowienia_dla_punktu), 
                        'Ilość w paletach': sum(zamowienia_dict[key][4] for key in zamowienia_dla_punktu),  
                        'GEOOKOORDYNATY': koord[2:22],
                        'Szerokość': lati,
                        'Długość': long,
                        'Data przejazdu' : date_analysis,
                        'Koszt' : koszt * (sum(zamowienia_dict[key][2] for key in zamowienia_dla_punktu))/(masa) ,
                        'Typ trasy': 'ZRODLO-KLIENT'
                        
                    }

                    new_rows_z.append(new_row_z)
                
        #print(new_rows_x)
###############################################################Export do Excel#################################################
        
        #timestr = time.strftime("%Y%m%d-%H%M%S")
        
        workbook_1 = xlsxwriter.Workbook(o_file+ "_"+date_analysis+'('+timestr+').xlsx')        
        
        format_licz = workbook_1.add_format({'num_format':'###0'})
        format_licz1 = workbook_1.add_format({'num_format':'0.00'})
                    
        worksheet_1 = workbook_1.add_worksheet("Optymalne trasy")
        worksheet_1.write(0,0,'Nr. Trasy')
        worksheet_1.write(0,1,'Źródło trasy')
        worksheet_1.write(0,2,'Punkt na trasie')
        worksheet_1.write(0,3,'Cel trasy')
        worksheet_1.write(0,4,'Nr. dostawy')
        worksheet_1.write(0,5,'Punkt wysyłkowy')
        worksheet_1.write(0,6,'Klient')
        worksheet_1.write(0,7,'MP')
        worksheet_1.write(0,8,'Waga [kg]')
        worksheet_1.write(0,9,'Przejechane KM')
        worksheet_1.write(0,10,'Koszt [PLN]')
        worksheet_1.write(0,11,'Koszt [PLN/kg]')
        worksheet_1.write(0,12,'GEOOKOORDYNATY')
        worksheet_1.write(0,13,'Szerokość')
        worksheet_1.write(0,14,'Długość')
        worksheet_1.write(0,15,'Data przejazdu')
        worksheet_1.write(0,16,'Typ trasy')

        worksheet_1.write(0,17, 'Wersja oprogramowania')
        worksheet_1.write(1,17, app_name)
        
        

            
        ka = 0

        len1 = len(new_rows_x)
        len2 = len(new_rows_z)
        len3 = len(new_rows_y)
            
        for i in range(len1+len2+len3):
            if i < len3:
                j = i
                nr_trasy = new_rows_y[j]['Nr. Trasy']
                if new_rows_y[j]['Źródło'] == "POSREDNIEPRUSZEWO":
                    mag_p = "MC NIEPRUSZEWO"
                else:
                    mag_p = new_rows_y[j]['Źródło']
                punkt_mp=new_rows_y[j]['Punkt na trasie']
                if new_rows_y[j]['Nazwa punktu'] == "POSREDNIEPRUSZEWO":
                    klient_mp = "MC NIEPRUSZEWO"
                else:
                    klient_mp=new_rows_y[j]['Nazwa punktu']  
                km_mp=new_rows_y[j]['Przejechane KM']
                #kg_mp=new_rows_y[j]['Ilość w kg']
                #palety_mp=new_rows_y[j]['Ilość w paletach']

                kg_mp=new_rows_y[j]['Ilość w kg']
                palety_mp=new_rows_y[j]['Ilość w paletach']

                geo_mp=new_rows_y[j]['GEOOKOORDYNATY']
                lati_mp=new_rows_y[j]['Szerokość']
                long_mp=new_rows_y[j]['Długość']
                date_mp=new_rows_y[j]['Data przejazdu']
                route_type_mp=new_rows_y[j]['Typ trasy']
                koszt_mp = new_rows_y[j]['Koszt']
                
                new_rows = new_rows_y[j]['Zamówienia']
            
            elif i < len3 + len1:
                j= i - len3
                
                nr_trasy = new_rows_x[j]['Nr. Trasy']
                
                if new_rows_x[j]['Źródło'] == "POSREDNIEPRUSZEWO":
                    mag_p = "MC NIEPRUSZEWO"
                else:
                    mag_p = new_rows_x[j]['Źródło']
                punkt_mp=new_rows_x[j]['Punkt na trasie']        
                if new_rows_x[j]['Nazwa punktu'] == "POSREDNIEPRUSZEWO":
                    klient_mp = "MC NIEPRUSZEWO"
                else:
                    klient_mp=new_rows_x[j]['Nazwa punktu']  
                km_mp=new_rows_x[j]['Przejechane KM']
                kg_mp=new_rows_x[j]['Ilość w kg']
                palety_mp=new_rows_x[j]['Ilość w paletach']
                geo_mp=new_rows_x[j]['GEOOKOORDYNATY']
                lati_mp=new_rows_x[j]['Szerokość']
                long_mp=new_rows_x[j]['Długość']
                date_mp=new_rows_x[j]['Data przejazdu']
                route_type_mp=new_rows_x[j]['Typ trasy']
                koszt_mp = new_rows_x[j]['Koszt']

                new_rows = new_rows_x[j]['Zamówienia']
            
            else:
                j= i - len3 - len1
                
                nr_trasy = new_rows_z[j]['Nr. Trasy']
                if new_rows_z[j]['Źródło'] == "POSREDNIEPRUSZEWO":
                    mag_p = "MC NIEPRUSZEWO"
                else:
                    mag_p = new_rows_z[j]['Źródło']
                punkt_mp=new_rows_z[j]['Punkt na trasie']
                if new_rows_z[j]['Nazwa punktu'] == "POSREDNIEPRUSZEWO":
                    klient_mp = "MC NIEPRUSZEWO"
                else:
                    klient_mp=new_rows_z[j]['Nazwa punktu']  
                km_mp=new_rows_z[j]['Przejechane KM']
                kg_mp=new_rows_z[j]['Ilość w kg']
                palety_mp=new_rows_z[j]['Ilość w paletach']
                geo_mp=new_rows_z[j]['GEOOKOORDYNATY']
                lati_mp=new_rows_z[j]['Szerokość']
                long_mp=new_rows_z[j]['Długość']
                date_mp=new_rows_z[j]['Data przejazdu']
                route_type_mp=new_rows_z[j]['Typ trasy']
                koszt_mp = new_rows_z[j]['Koszt']

                new_rows = new_rows_z[j]['Zamówienia']
            

    
            #len2 = len(new_rows_z[i]['Zamówienia'])
             
            for zam in range(1+len(new_rows)):
                #if sh == sh_names[0]:
                #   zam_mp = new_rows_x[i]['Zamówienia'][zam]
                #elif sh == sh_names[1]
                #    zam_mp = new_rows_z[i]['Zamówienia'][zam]
                
                if len(new_rows)!=0:
                    if zam !=len(new_rows):
                        zam_mp = new_rows[zam]
                    else:
                        break
                else:
                    zam_mp = 0
                
                worksheet_1.write(ka+1,0,nr_trasy)
                worksheet_1.write(ka+1,1,mag_p)
                worksheet_1.write(ka+1,2,punkt_mp)
                worksheet_1.write(ka+1,3,klient_mp)
                worksheet_1.write(ka+1,4,float(zam_mp),format_licz)
                worksheet_1.write(ka+1,9,float(km_mp),format_licz)
                #worksheet_1.write(ka+1,6,float(kg_mp),format_licz)
                
                #worksheet_1.write(ka+1,7,float(palety_mp),format_licz)
                worksheet_1.write(ka+1,12,str(geo_mp))
                worksheet_1.write(ka+1,13,str(lati_mp))
                worksheet_1.write(ka+1,14,str(long_mp))
                worksheet_1.write(ka+1,15,str(date_mp))
                worksheet_1.write(ka+1,16,str(route_type_mp))
                if zam_mp !=0:
                    worksheet_1.write(ka+1,8,float(zamowienia_dict[zam_mp][2]),format_licz)
                    worksheet_1.write(ka+1,7,float(zamowienia_dict[zam_mp][4]),format_licz)
                    worksheet_1.write(ka+1,5,str(zamowienia_dict[zam_mp][1]))
                    worksheet_1.write(ka+1,6,zamowienia_dict[zam_mp][0])
                    if kg_mp !=0:
                        worksheet_1.write(ka+1,10,(koszt_mp/(float(kg_mp)))*zamowienia_dict[zam_mp][2], format_licz1)
                        worksheet_1.write(ka+1,11,(koszt_mp/(float(kg_mp))), format_licz1)
                
                
                ka+=1
                #zam_mp.clear()
            
        if report == 1:
            print(mag_p)
            #print(tras_mp)
            print(geo_mp)

        worksheet_1 = workbook_1.add_worksheet("Parametry analizy")

        # Nazwy charakterystyk
        worksheet_1.write(0,0,'PARAMETRY ANALIZY')
        worksheet_1.write(1,0,'Auta')
        worksheet_1.write(2,0,'Liczba aut dostępna dla magazynu')
        worksheet_1.write(3,0,'Ładowność auta [kg]')
        worksheet_1.write(4,0,'Pojemność auta [palety]')
        worksheet_1.write(5,0,'Minimalne wypełnienie aut	')
        worksheet_1.write(6,0,'Trasy')
        worksheet_1.write(7,0,'Max. dystans w dostawie bezpośredniej [km]')
        worksheet_1.write(8,0,'Max. dystans w dostawie pośredniej [km]')
        worksheet_1.write(9,0,'Wypełnienie auta --> transport bezpośredni')
        worksheet_1.write(10,0,'Koszty')
        worksheet_1.write(11,0,'Transport')
        worksheet_1.write(12,0,'Koszt transportu [zł/km]')
        worksheet_1.write(13,0,'Fracht')
        worksheet_1.write(14,0,'MC Niepruszewo --> ESA (Říčany) [EUR]')
        worksheet_1.write(15,0,'MC Bytom --> MECOM (Lucenec) [PLN]')
        worksheet_1.write(16,0,'Przeładunek')
        worksheet_1.write(17,0,'MC Niepruszewo [PLN/kg]')
        worksheet_1.write(18,0,'MC Bytom [PLN/paleta]')
        worksheet_1.write(19,0,'Operatorzy logistyczni')
        worksheet_1.write(20,0,'Koszt ESA [EUR/kg]')
        worksheet_1.write(21,0,'Koszt MECOM-->SK [EUR/paleta]')
        worksheet_1.write(22,0,'Koszt MECOM-->HU [EUR/paleta]')
        worksheet_1.write(23,0,'Pozostałe')
        worksheet_1.write(24,0,'Kurs wymiany EUR/PLN')
        worksheet_1.write(25,0,'Miejsca paletowe')
        worksheet_1.write(26,0,'Wędlin na palecie [kg]')
        worksheet_1.write(27,0,'Mięsa na palecie [kg]')
        worksheet_1.write(28,0,'Pozostałe produkcty na paleciee [kg]')
        
        worksheet_1.write(29,0, 'Wersja oprogramowania')
    

        
       # Wartości charakterystyk
        worksheet_1.write(0,1,'Wartość')
        #worksheet_1.write(1,1,'Auta')
        worksheet_1.write(2,1,k)
        worksheet_1.write(3,1,maximum)
        worksheet_1.write(4,1,maxpalet)
        worksheet_1.write(5,1,min_wyp)
        #worksheet_1.write(6,1,'Trasy')
        worksheet_1.write(7,1,MAX_DISTANCE_BEZP)
        worksheet_1.write(8,1,MAX_DISTANCE_POS)
        worksheet_1.write(9,1,wymuszenie)
        #worksheet_1.write(10,1,'Koszty')
        #worksheet_1.write(11,1,'Transport')
        worksheet_1.write(12,1,koszt_km)
        #worksheet_1.write(13,1,'Fracht')
        worksheet_1.write(14,1,koszt_fracht_niep)
        worksheet_1.write(15,1,koszt_fracht_bytom)
        #worksheet_1.write(16,1,'Przeładunek')
        worksheet_1.write(17,1,przel_niep_kg)
        worksheet_1.write(18,1,przel_bytom_pal)
        #worksheet_1.write(19,1,'Operatorzy logistyczni')
        worksheet_1.write(20,1,kg_op_niepruszewo)
        worksheet_1.write(21,1,mecom_sk)
        worksheet_1.write(22,1,mecom_hu)
        
        #worksheet_1.write(23,1,'Pozostałe')
        worksheet_1.write(24,1,kurs_euro)
        #worksheet_1.write(25,1,'Miejsca paletowe')
        worksheet_1.write(26,1,dfp['Value'][40])
        worksheet_1.write(27,1,dfp['Value'][41])
        worksheet_1.write(28,1,dfp['Value'][42])
        
        worksheet_1.write(29,1, app_name)
        
        workbook_1.close()
    
        df_trasy_z_expanded = pd.DataFrame(new_rows_z)

        df_trasy_expanded = pd.concat([df_trasy_y_expanded, df_trasy_x_expanded, df_trasy_z_expanded], ignore_index=True)

        #df_trasy_expanded['Źródło'] = df_trasy_expanded['Źródło'].replace('POSREDNIEPRUSZEWO', 'MC Niepruszewo')
        #df_trasy_expanded['Nazwa punktu'] = df_trasy_expanded['Nazwa punktu'].replace('POSREDNIEPRUSZEWO', 'MC Niepruszewo')

        df_trasy_expanded1 = df_trasy_expanded.replace('POSREDNIEPRUSZEWO', 'MC NIEPRUSZEWO')
        
        #i = df[((df.Źródło == 'MC NIEPRUSZEWO') &( df.Nazwa == 15))].index

        #display(df_trasy_y_expanded)
        #display(df_trasy_z_expanded)
        #display(df_trasy_x_expanded)
        display(df_trasy_expanded1)
        
        # zmiana kolejności wypisywania df'ow do xlsxa
        #with pd.ExcelWriter(o_file +'_DF'+date_analysis+'('+timestr+').xlsx') as dane_wy:
        #    try:
        #        df_trasy_y_expanded.to_excel(dane_wy,sheet_name='Mag_zrod_pos')
        #    except:
        #        print("Brak danych")
        #    try:
        #        df_trasy_z_expanded.to_excel(dane_wy,sheet_name='Mag_zrodlowy')
        #    except:
        #        print("Brak danych")
        #    try:
        #        df_trasy_x_expanded.to_excel(dane_wy,sheet_name='Mag_posredni')
        #    except:
        #        print("Brak danych")
        
        with  pd.ExcelWriter(o_file + "_" +date_analysis+'('+timestr+').xlsx',engine='openpyxl', mode='a') as workbook_1:
        #with pd.ExcelWriter(o_file + '_DF'+date_analysis+'('+timestr+').xlsx') as dane_wy:
            try:
                df_trasy_expanded1.to_excel(workbook_1,sheet_name='Optymalne_trasy_DF')
            except:
                print("Brak danych")
        with  pd.ExcelWriter(o_file + "_" +date_analysis+'('+timestr+').xlsx',engine='openpyxl', mode='a') as workbook_1:
            try:
                dfp.to_excel(workbook_1,sheet_name='Parametry_analizyDF')
            except:
                print("Brak danych")

print('+++++++++++++++++++++++++++++++++++++++++++')
print('Aplikacja transExpApp FINISHED!')
print('+++++++++++++++++++++++++++++++++++++++++++')  